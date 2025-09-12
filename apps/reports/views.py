from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import View, TemplateView
from django.db.models import Sum, Count, Avg, Q, F, Max, Min, Value, Case, When, DecimalField
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek, TruncYear, Extract
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import io
import json
from django.db import models

# Third party imports
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.units import inch
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Model imports
from apps.services.models import Service, ServiceCategory, ServiceOrder, ServiceOrderItem, ServicePackage
from apps.employees.models import Employee, Department, Attendance, PerformanceReview
from apps.customers.models import Customer, Vehicle, LoyaltyProgram
from apps.payments.models import Payment, PaymentMethod
from apps.inventory.models import InventoryItem, StockMovement, InventoryCategory
from apps.expenses.models import Expense, ExpenseCategory, Vendor
from apps.suppliers.models import Supplier, PurchaseOrder
from apps.subscriptions.models import Subscription
from apps.core.decorators import business_required

@method_decorator([login_required, business_required], name='dispatch')
class ReportsView(TemplateView):
    """Comprehensive focused reports system showing specific report data"""
    template_name = 'reports/reports.html'
    
    def get(self, request, *args, **kwargs):
        # Check if this is an export request
        export_format = request.GET.get('export')
        if export_format in ['pdf', 'excel']:
            return self._handle_export(request, export_format)
        
        # Otherwise, render the normal template
        return super().get(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get report type from URL parameters
        report_type = self.request.GET.get('report_type', 'business_overview')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        page = int(self.request.GET.get('page', 1))
        
        # Set default date range if not provided
        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        else:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                # If date parsing fails, use default range
                end_date = timezone.now().date()
                start_date = end_date - timedelta(days=30)
        
        # Get report-specific data
        report_data = self._get_report_data(report_type, start_date, end_date, page)
        
        context.update({
            'selected_report_type': report_type,
            'start_date': start_date,
            'end_date': end_date,
            'current_page': page,
            'report_data': report_data,
            'available_reports': self._get_available_reports(),
            'pdf_available': PDF_AVAILABLE,
            'excel_available': EXCEL_AVAILABLE,
            'employees_list': self._get_employees_list() if report_type == 'individual_employee' else [],
            'selected_employee_id': self.request.GET.get('employee_id', ''),
        })
        
        return context
    
    def _get_employees_list(self):
        """Get list of employees for dropdown"""
        try:
            employees = Employee.objects.filter(is_active=True).select_related('department')
            return [(emp.id, f"{emp.full_name} - {emp.department.name if emp.department else 'No Dept'}") 
                   for emp in employees]
        except Exception as e:
            return [('', f'Error loading employees: {str(e)}')]
    
    def _get_datetime_range(self, start_date, end_date):
        """Convert date objects to proper datetime range for filtering"""
        from django.utils import timezone
        start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
        return start_datetime, end_datetime
    
    def _get_available_reports(self):
        """Get list of available report types"""
        return [
            ('business_overview', 'Business Overview'),
            ('inventory', 'Inventory Management'),
            ('services', 'Services Performance'),
            ('customers', 'Customer Analysis'),
            ('employees', 'Employee Performance'),
            ('payments', 'Payment Transactions'),
            ('expenses', 'Expense Analysis'),
            ('individual_employee', 'Individual Employee'),
            ('financial_summary', 'Financial Summary'),
            ('daily_summary', 'Daily Summary'),
            ('weekly_summary', 'Weekly Summary'),
            ('monthly_summary', 'Monthly Summary'),
            ('sales_analysis', 'Sales Analysis'),
            ('customer_analytics', 'Customer Analytics'),
            ('supplier_report', 'Supplier Performance'),
            ('vehicle_analysis', 'Vehicle Analysis'),
            ('service_packages', 'Service Packages'),
            ('loyalty_program', 'Loyalty Program'),
            ('attendance_report', 'Employee Attendance'),
            ('department_performance', 'Department Performance'),
            ('subscription_analysis', 'Subscription Analysis'),
            ('balance_sheet', 'Balance Sheet'),
            ('refunds_report', 'Refunds & Returns'),
        ]
    
    def _get_report_data(self, report_type, start_date, end_date, page):
        """Get data specific to the report type"""
        data = {
            'title': self._get_report_title(report_type),
            'type': report_type,
            'period': f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            'items': [],
            'summary': {},
            'pagination': {},
        }
        
        try:
            if report_type == 'inventory':
                data.update(self._get_inventory_data(start_date, end_date, page))
            elif report_type == 'services':
                data.update(self._get_services_data(start_date, end_date, page))
            elif report_type == 'customers':
                data.update(self._get_customers_data(start_date, end_date, page))
            elif report_type == 'employees':
                data.update(self._get_employees_data(start_date, end_date, page))
            elif report_type == 'payments':
                data.update(self._get_payments_data(start_date, end_date, page))
            elif report_type == 'expenses':
                data.update(self._get_expenses_data(start_date, end_date, page))
            elif report_type == 'individual_employee':
                employee_id = self.request.GET.get('employee_id')
                data.update(self._get_individual_employee_data(employee_id, start_date, end_date, page))
            elif report_type == 'financial_summary':
                data.update(self._get_financial_summary_data(start_date, end_date, page))
            elif report_type == 'daily_summary':
                data.update(self._get_daily_summary_data(start_date, end_date, page))
            elif report_type == 'weekly_summary':
                data.update(self._get_weekly_summary_data(start_date, end_date, page))
            elif report_type == 'monthly_summary':
                data.update(self._get_monthly_summary_data(start_date, end_date, page))
            elif report_type == 'sales_analysis':
                data.update(self._get_sales_analysis_data(start_date, end_date, page))
            elif report_type == 'customer_analytics':
                data.update(self._get_customer_analytics_data(start_date, end_date, page))
            elif report_type == 'supplier_report':
                data.update(self._get_supplier_report_data(start_date, end_date, page))
            elif report_type == 'vehicle_analysis':
                data.update(self._get_vehicle_analysis_data(start_date, end_date, page))
            elif report_type == 'service_packages':
                data.update(self._get_service_packages_data(start_date, end_date, page))
            elif report_type == 'loyalty_program':
                data.update(self._get_loyalty_program_data(start_date, end_date, page))
            elif report_type == 'attendance_report':
                data.update(self._get_attendance_report_data(start_date, end_date, page))
            elif report_type == 'department_performance':
                data.update(self._get_department_performance_data(start_date, end_date, page))
            elif report_type == 'subscription_analysis':
                data.update(self._get_subscription_analysis_data(start_date, end_date, page))
            elif report_type == 'balance_sheet':
                data.update(self._get_balance_sheet_data(start_date, end_date, page))
            elif report_type == 'refunds_report':
                data.update(self._get_refunds_report_data(start_date, end_date, page))
            else:  # business_overview
                data.update(self._get_business_overview_data(start_date, end_date, page))
                
        except Exception as e:
            data['error'] = f"Error generating {report_type} report: {str(e)}"
            
        return data
    
    def _get_report_title(self, report_type):
        """Get the title for a specific report type"""
        titles = {
            'business_overview': 'Business Overview Report',
            'inventory': 'Inventory Management Report',
            'services': 'Services Performance Report',
            'customers': 'Customer Analysis Report',
            'employees': 'Employee Performance Report',
            'payments': 'Payment Transactions Report',
            'expenses': 'Expense Analysis Report',
            'individual_employee': 'Individual Employee Report',
            'financial_summary': 'Financial Summary Report',
            'daily_summary': 'Daily Summary Report',
            'weekly_summary': 'Weekly Summary Report',
            'monthly_summary': 'Monthly Summary Report',
            'sales_analysis': 'Sales Analysis Report',
            'customer_analytics': 'Customer Analytics Report',
            'supplier_report': 'Supplier Performance Report',
            'vehicle_analysis': 'Vehicle Analysis Report',
            'service_packages': 'Service Packages Report',
            'loyalty_program': 'Loyalty Program Report',
            'attendance_report': 'Employee Attendance Report',
            'department_performance': 'Department Performance Report',
            'subscription_analysis': 'Subscription Analysis Report',
            'balance_sheet': 'Balance Sheet Report',
            'refunds_report': 'Refunds & Returns Report',
        }
        return titles.get(report_type, 'Business Report')
    
    # ============== CORE REPORT DATA METHODS ==============
    
    def _get_business_overview_data(self, start_date, end_date, page):
        """Get business overview data based on service orders created in the period"""
        from django.core.paginator import Paginator
        
        try:
            # Convert dates to datetime for proper filtering
            start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
            
            # Get service orders created in the date range
            service_orders = ServiceOrder.objects.select_related(
                'customer', 
                'assigned_attendant', 
                'vehicle'
            ).filter(
                created_at__range=[start_datetime, end_datetime]
            ).order_by('-created_at')
            
            # Paginate service orders
            paginator = Paginator(service_orders, 20)
            orders_page = paginator.get_page(page)
            
            # Get payments for these orders (regardless of when payment was made)
            orders_with_payments = service_orders.filter(
                payments__status__in=['completed', 'verified']
            ).exclude(
                payments__payment_type='refund'
            ).distinct()
            
            # Calculate summary metrics based on orders created in the period
            total_orders = service_orders.count()
            paid_orders_count = orders_with_payments.count()
            
            # Revenue from payments for orders created in this period
            total_revenue = Payment.objects.filter(
                service_order__in=service_orders,
                status__in=['completed', 'verified']
            ).exclude(
                payment_type='refund'
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            # Track refunds separately for orders created in this period
            total_refunds = Payment.objects.filter(
                service_order__in=service_orders,
                payment_type='refund',
                status__in=['completed', 'verified']
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            # Also get refunds from PaymentRefund model
            from apps.payments.models import PaymentRefund
            refund_records = PaymentRefund.objects.filter(
                original_payment__service_order__in=service_orders,
                status='completed'
            ).aggregate(Sum('amount'))['amount__sum'] or 0
            
            # Total refunds = refund payments + refund records
            total_refunds_combined = total_refunds + refund_records
            
            # Net revenue = revenue - refunds
            net_revenue = total_revenue - total_refunds_combined
            
            # Orders not yet paid
            pending_orders = total_orders - paid_orders_count
            
            # Total order value (regardless of payment status)
            total_order_value = service_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
            
            avg_order_value = total_order_value / max(total_orders, 1)
            payment_completion_rate = (paid_orders_count / max(total_orders, 1)) * 100
            
            # Debug information
            debug_info = {
                'orders_created_count': total_orders,
                'paid_orders_count': paid_orders_count,
                'revenue_from_paid_orders': total_revenue,
                'pending_orders': pending_orders,
                'total_order_value': total_order_value,
                'date_range': f"{start_date} to {end_date}",
                'datetime_range': f"{start_datetime} to {end_datetime}",
                'payment_completion_rate': payment_completion_rate,
                'tenant_id': getattr(self.request.tenant, 'id', 'Unknown') if hasattr(self.request, 'tenant') else 'No tenant'
            }
            
            return {
                'items': orders_page,
                'summary': {
                    'total_orders': total_orders,
                    'total_revenue': total_revenue,
                    'total_refunds': total_refunds_combined,
                    'net_revenue': net_revenue,
                    'paid_orders': paid_orders_count,
                    'pending_orders': pending_orders,
                    'avg_order_value': avg_order_value,
                    'payment_completion_rate': payment_completion_rate,
                    'refund_rate': (total_refunds_combined / max(total_revenue, 1)) * 100,
                },
                'pagination': self._get_pagination_data(orders_page, paginator),
                'debug_info': debug_info
            }
        except Exception as e:
            return {
                'items': [],
                'summary': {
                    'total_orders': 0,
                    'total_revenue': 0,
                    'total_refunds': 0,
                    'net_revenue': 0,
                    'paid_orders': 0,
                    'pending_orders': 0,
                    'avg_order_value': 0,
                    'payment_completion_rate': 0,
                    'refund_rate': 0,
                },
                'pagination': {},
                'error': f"Error generating business overview: {str(e)}",
                'debug_info': {'error': str(e)}
            }
    
    def _get_inventory_data(self, start_date, end_date, page):
        """Get inventory-specific data"""
        from django.core.paginator import Paginator
        
        try:
            # Convert dates to datetime for proper filtering
            start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
            
            # Get inventory items with movements in the date range
            items = InventoryItem.objects.select_related('category', 'unit').annotate(
                total_in=Sum('stock_movements__quantity', 
                            filter=Q(stock_movements__created_at__range=[start_datetime, end_datetime],
                                   stock_movements__movement_type='in')),
                total_out=Sum('stock_movements__quantity', 
                             filter=Q(stock_movements__created_at__range=[start_datetime, end_datetime],
                                    stock_movements__movement_type='out')),
                movement_count=Count('stock_movements', 
                                   filter=Q(stock_movements__created_at__range=[start_datetime, end_datetime]))
            ).order_by('-movement_count', 'name')
            
            # Paginate items
            paginator = Paginator(items, 20)
            items_page = paginator.get_page(page)
            
            # Calculate summary
            total_items = items.count()
            total_value = sum(item.current_stock * (item.unit_cost or 0) for item in items)
            low_stock_items = items.filter(current_stock__lte=F('minimum_stock_level')).count()
            
            return {
                'items': items_page,
                'summary': {
                    'total_items': total_items,
                    'total_value': total_value,
                    'low_stock_items': low_stock_items,
                    'debug_info': f"Items with movements: {items.filter(movement_count__gt=0).count()}",
                },
                'pagination': self._get_pagination_data(items_page, paginator)
            }
        except Exception as e:
            return {
                'items': [],
                'summary': {},
                'pagination': {},
                'error': f'Inventory report error: {str(e)}'
            }
    
    def _get_services_data(self, start_date, end_date, page):
        """Get services-specific data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        try:
            # Convert dates to datetime for proper filtering
            start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
            
            # Get services with performance metrics based on orders created in the period
            services = Service.objects.select_related('category').annotate(
                # Count orders created in the period
                orders_count=Count('order_items__order',
                                 filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime])),
                # Total order value for orders created in the period
                total_order_value=Sum('order_items__total_price',
                                    filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime])),
                # Revenue from payments for orders created in this period (regardless of when payment was made)
                actual_revenue=Sum('order_items__order__payments__amount',
                                 filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime],
                                         order_items__order__payments__status__in=['completed', 'verified']) &
                                         ~Q(order_items__order__payments__payment_type='refund')),
                # Count of paid orders created in the period
                paid_orders_count=Count('order_items__order__payments',
                                       distinct=True,
                                       filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime],
                                               order_items__order__payments__status__in=['completed', 'verified']) & 
                                               ~Q(order_items__order__payments__payment_type='refund')),
                avg_rating=Avg('order_items__rating',
                             filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime]))
            ).filter(orders_count__gt=0).order_by('-actual_revenue', 'name')
            
            # Paginate services
            paginator = Paginator(services, 20)
            services_page = paginator.get_page(page)
            
            # Calculate summary
            total_services = services.count()
            total_orders = sum(s.orders_count or 0 for s in services)
            total_paid_orders = sum(s.paid_orders_count or 0 for s in services)
            total_order_value = sum(s.total_order_value or 0 for s in services)
            total_revenue = sum(s.actual_revenue or 0 for s in services)
            
            return {
                'items': services_page,
                'summary': {
                    'total_services': total_services,
                    'total_orders': total_orders,
                    'total_paid_orders': total_paid_orders,
                    'total_order_value': total_order_value,
                    'total_revenue': total_revenue,
                    'avg_revenue_per_service': total_revenue / max(total_services, 1),
                    'avg_order_value': total_order_value / max(total_orders, 1),
                    'payment_rate': (total_paid_orders / max(total_orders, 1)) * 100,
                },
                'pagination': self._get_pagination_data(services_page, paginator)
            }
        except Exception as e:
            return {
                'items': [],
                'summary': {},
                'pagination': {},
                'error': f'Services report error: {str(e)}'
            }
    
    def _get_customers_data(self, start_date, end_date, page):
        """Get customers-specific data based on actual payments (excluding refunds)"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        # Get customers with their payment activity in the date range (excluding refunds)
        customers = Customer.objects.annotate(
            # Count payments made in the period (excluding refunds)
            payments_count=Count('payments',
                               filter=Q(payments__completed_at__range=[start_datetime, end_datetime],
                                       payments__status__in=['completed', 'verified']) &
                                       ~Q(payments__payment_type='refund')),
            # Total amount actually paid (excluding refunds)
            customer_total_spent=Sum('payments__amount',
                          filter=Q(payments__completed_at__range=[start_datetime, end_datetime],
                                 payments__status__in=['completed', 'verified']) &
                                 ~Q(payments__payment_type='refund')),
            # Last payment date (excluding refunds)
            last_payment=Max('payments__completed_at',
                           filter=Q(payments__completed_at__range=[start_datetime, end_datetime],
                                   payments__status__in=['completed', 'verified']) &
                                   ~Q(payments__payment_type='refund')),
            # Count of orders with payments (excluding refunds)
            paid_orders_count=Count('service_orders__payments',
                                  filter=Q(service_orders__payments__completed_at__range=[start_datetime, end_datetime],
                                          service_orders__payments__status__in=['completed', 'verified']) &
                                          ~Q(service_orders__payments__payment_type='refund'))
        ).order_by('-customer_total_spent')
        
        # Paginate customers
        paginator = Paginator(customers, 20)
        customers_page = paginator.get_page(page)
        
        # Calculate summary
        total_customers = customers.count()
        active_customers = customers.filter(payments_count__gt=0).count()
        total_revenue = sum(c.customer_total_spent or 0 for c in customers)
        total_payments = sum(c.payments_count or 0 for c in customers)
        avg_payment_value = total_revenue / max(total_payments, 1)
        
        return {
            'items': customers_page,
            'summary': {
                'total_customers': total_customers,
                'active_customers': active_customers,
                'total_revenue': total_revenue,
                'total_payments': total_payments,
                'avg_payment_value': avg_payment_value,
                'customer_retention_rate': (active_customers / max(total_customers, 1)) * 100,
            },
            'pagination': self._get_pagination_data(customers_page, paginator)
        }
    
    def _get_employees_data(self, start_date, end_date, page):
        """Get employees-specific data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        # Get employees with their performance metrics based on orders created in the period
        employees = Employee.objects.select_related('department').annotate(
            # Total orders assigned to employee that were created in the period
            total_orders_assigned=Count('assigned_orders',
                                      filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime])),
            # Revenue from payments for orders created in this period (regardless of when payment was made)
            revenue_generated=Sum('assigned_orders__payments__amount',
                                filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime],
                                        assigned_orders__payments__status__in=['completed', 'verified']) &
                                        ~Q(assigned_orders__payments__payment_type='refund')),
            # Orders created in period that have been paid
            paid_orders_handled=Count('assigned_orders__payments',
                                    distinct=True,
                                    filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime],
                                            assigned_orders__payments__status__in=['completed', 'verified']) &
                                            ~Q(assigned_orders__payments__payment_type='refund')),
            # Total order value for orders created in period
            total_order_value=Sum('assigned_orders__total_amount',
                                filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime])),
            # Performance metrics
            payment_conversion_rate=Case(
                When(total_orders_assigned__gt=0, 
                     then=(F('paid_orders_handled') * 100.0) / F('total_orders_assigned')),
                default=0,
                output_field=DecimalField(max_digits=5, decimal_places=2)
            )
        ).filter(total_orders_assigned__gt=0).order_by('-revenue_generated')
        
        # Paginate employees
        paginator = Paginator(employees, 20)
        employees_page = paginator.get_page(page)
        
        # Calculate summary
        total_employees = employees.count()
        total_orders_assigned = sum(e.total_orders_assigned or 0 for e in employees)
        total_paid_orders = sum(e.paid_orders_handled or 0 for e in employees)
        total_revenue = sum(e.revenue_generated or 0 for e in employees)
        total_order_value = sum(e.total_order_value or 0 for e in employees)
        avg_revenue_per_employee = total_revenue / max(total_employees, 1)
        avg_orders_per_employee = total_orders_assigned / max(total_employees, 1)
        
        return {
            'items': employees_page,
            'summary': {
                'total_employees': total_employees,
                'total_orders_assigned': total_orders_assigned,
                'total_paid_orders': total_paid_orders,
                'total_revenue': total_revenue,
                'total_order_value': total_order_value,
                'avg_revenue_per_employee': avg_revenue_per_employee,
                'avg_orders_per_employee': avg_orders_per_employee,
                'payment_rate': (total_paid_orders / max(total_orders_assigned, 1)) * 100,
            },
            'pagination': self._get_pagination_data(employees_page, paginator)
        }
    
    def _get_payments_data(self, start_date, end_date, page):
        """Get payments-specific data for services ordered in the date range"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        # Get payments for service orders created in the date range (not payments made in date range)
        payments = Payment.objects.select_related('payment_method', 'service_order', 'customer').filter(
            service_order__created_at__range=[start_datetime, end_datetime],
            status__in=['completed', 'verified']
        ).exclude(
            payment_type='refund'
        ).order_by('-completed_at', '-created_at')
        
        # Paginate payments
        paginator = Paginator(payments, 20)
        payments_page = paginator.get_page(page)
        
        # Calculate summary
        total_payments = payments.aggregate(Sum('amount'))['amount__sum'] or 0
        payment_count = payments.count()
        
        # Payment method breakdown
        method_breakdown = payments.values('payment_method__name').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')
        
        return {
            'items': payments_page,
            'summary': {
                'total_payments': total_payments,
                'payment_count': payment_count,
                'avg_payment': total_payments / max(payment_count, 1),
                'method_breakdown': list(method_breakdown),
            },
            'pagination': self._get_pagination_data(payments_page, paginator)
        }
    
    def _get_expenses_data(self, start_date, end_date, page):
        """Get expenses-specific data - only approved expenses"""
        from django.core.paginator import Paginator
        
        # Get only approved expenses in the date range
        expenses = Expense.objects.select_related('category', 'vendor').filter(
            expense_date__range=[start_date, end_date],
            status='approved'  # Only show approved expenses
        ).order_by('-expense_date')
        
        # Paginate expenses
        paginator = Paginator(expenses, 20)
        expenses_page = paginator.get_page(page)
        
        # Calculate summary
        total_expenses = expenses.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        expense_count = expenses.count()
        
        # Category breakdown
        category_breakdown = expenses.values('category__name').annotate(
            total=Sum('total_amount'),
            count=Count('id')
        ).order_by('-total')
        
        return {
            'items': expenses_page,
            'summary': {
                'total_expenses': total_expenses,
                'expense_count': expense_count,
                'avg_expense': total_expenses / max(expense_count, 1),
                'category_breakdown': list(category_breakdown),
            },
            'pagination': self._get_pagination_data(expenses_page, paginator)
        }
    
    def _get_individual_employee_data(self, employee_id, start_date, end_date, page):
        """Get individual employee-specific data with detailed information"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        if not employee_id:
            return {'error': 'Employee ID is required'}
        
        try:
            employee = Employee.objects.select_related('department').get(id=employee_id)
        except Employee.DoesNotExist:
            return {'error': 'Employee not found'}
        
        # Get payments for orders created and handled by this employee in the date range (excluding refunds)
        employee_payments = Payment.objects.select_related(
            'service_order__customer', 
            'service_order__vehicle',
            'payment_method'
        ).filter(
            service_order__assigned_attendant=employee,
            service_order__created_at__range=[start_datetime, end_datetime],  # Orders created in period
            status__in=['completed', 'verified']
        ).exclude(
            payment_type='refund'  # Exclude refunds
        ).order_by('-completed_at')
        
        # Paginate payments
        paginator = Paginator(employee_payments, 20)
        payments_page = paginator.get_page(page)
        
        # Calculate employee-specific summary based on payments (excluding refunds)
        total_payments = employee_payments.count()
        total_revenue = employee_payments.aggregate(Sum('amount'))['amount__sum'] or 0
        avg_payment_value = total_revenue / max(total_payments, 1)
        
        # Get orders assigned (may not all be paid)
        all_orders_assigned = ServiceOrder.objects.filter(
            assigned_attendant=employee,
            created_at__range=[start_datetime, end_datetime]
        )
        total_orders_assigned = all_orders_assigned.count()
        
        # Payment conversion rate
        payment_conversion_rate = (total_payments / max(total_orders_assigned, 1)) * 100
        
        # Get attendance data if available
        try:
            from apps.employees.models import Attendance
            attendance_records = Attendance.objects.filter(
                employee=employee,
                date__range=[start_date, end_date]
            )
            total_days = attendance_records.count()
            present_days = attendance_records.filter(status='present').count()
            attendance_rate = (present_days / max(total_days, 1)) * 100 if total_days > 0 else 0
        except:
            attendance_rate = 0
            total_days = 0
            present_days = 0
        
        # Payment method breakdown (excluding refunds)
        payment_methods = employee_payments.values('payment_method__name').annotate(
            count=Count('id'),
            total_amount=Sum('amount')
        ).order_by('-total_amount')
        
        return {
            'employee': employee,
            'items': payments_page,
            'summary': {
                'total_payments': total_payments,
                'revenue_generated': total_revenue,
                'total_revenue': total_revenue,  # For consistency with template
                'avg_payment_value': avg_payment_value,
                'total_orders_assigned': total_orders_assigned,
                'payment_conversion_rate': payment_conversion_rate,
                'attendance_rate': attendance_rate,
                'total_attendance_days': total_days,
                'present_days': present_days,
                'absent_days': total_days - present_days,
            },
            'payment_methods': payment_methods,
            'employee_details': {
                'employee_id': getattr(employee, 'employee_id', f"EMP-{employee.id}"),
                'full_name': employee.full_name,
                'email': getattr(employee, 'email', 'N/A'),
                'phone_number': str(getattr(employee, 'phone', 'N/A')),  # Use phone field but name it phone_number for template
                'department': employee.department.name if employee.department else 'N/A',
                'position': getattr(employee, 'position', 'N/A'),
                'hire_date': getattr(employee, 'hire_date', 'N/A'),
                'is_active': employee.is_active,
            },
            'pagination': self._get_pagination_data(payments_page, paginator)
        }
    
    # ============== ADVANCED REPORT DATA METHODS ==============
    
    def _get_financial_summary_data(self, start_date, end_date, page):
        """Get financial summary data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        # Get service orders created in the period
        orders_in_period = ServiceOrder.objects.filter(
            created_at__range=[start_datetime, end_datetime]
        )
        
        # Get revenue from payments for orders created in this period (regardless of when payment was made)
        revenue = Payment.objects.filter(
            service_order__in=orders_in_period,
            status__in=['completed', 'verified']
        ).exclude(
            payment_type='refund'  # Exclude refunds from revenue
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Track refunds separately for orders created in this period
        refunds = Payment.objects.filter(
            service_order__in=orders_in_period,
            payment_type='refund',
            status__in=['completed', 'verified']
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Also get refunds from PaymentRefund model
        from apps.payments.models import PaymentRefund
        refund_records = PaymentRefund.objects.filter(
            original_payment__service_order__in=orders_in_period,
            status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Total refunds = refund payments + refund records
        total_refunds = refunds + refund_records
        
        # Net revenue = revenue - refunds
        net_revenue = revenue - total_refunds
        
        # Get only approved expenses for the period
        expenses = Expense.objects.filter(
            expense_date__range=[start_date, end_date],
            status='approved'  # Only approved expenses
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Get recent financial transactions
        transactions = []
        
        # Add successful payments for orders created in this period as positive transactions
        payments = Payment.objects.select_related('payment_method', 'service_order').filter(
            service_order__in=orders_in_period,
            status__in=['completed', 'verified']
        ).exclude(
            payment_type='refund'
        ).order_by('-completed_at')[:50]
        
        for payment in payments:
            transactions.append({
                'type': 'Revenue',
                'amount': payment.amount,
                'description': f"Payment for Order #{payment.service_order.order_number if payment.service_order else 'N/A'} (Created: {payment.service_order.created_at.date() if payment.service_order else 'N/A'})",
                'date': payment.completed_at.date() if hasattr(payment.completed_at, 'date') else payment.completed_at,
                'category': payment.payment_method.name if payment.payment_method else 'Unknown'
            })
        
        # Add refunds for orders created in this period as negative impact transactions
        refunds = Payment.objects.select_related('payment_method', 'service_order').filter(
            service_order__in=orders_in_period,
            status__in=['completed', 'verified'],
            payment_type='refund'
        ).order_by('-completed_at')[:25]
        
        for refund in refunds:
            transactions.append({
                'type': 'Refund',
                'amount': -refund.amount,  # Negative impact
                'description': f"Refund for Order #{refund.service_order.order_number if refund.service_order else 'N/A'} (Created: {refund.service_order.created_at.date() if refund.service_order else 'N/A'})",
                'date': refund.completed_at.date() if hasattr(refund.completed_at, 'date') else refund.completed_at,
                'category': refund.payment_method.name if refund.payment_method else 'Unknown'
            })
        
        # Add only approved expenses as negative transactions
        expense_items = Expense.objects.select_related('category', 'vendor').filter(
            expense_date__range=[start_date, end_date],
            status='approved'  # Only approved expenses
        ).order_by('-expense_date')[:25]
        
        for expense in expense_items:
            transactions.append({
                'type': 'Expense',
                'amount': -expense.total_amount,
                'description': expense.description or f"Expense - {expense.category.name if expense.category else 'General'}",
                'date': expense.expense_date,
                'category': expense.category.name if expense.category else 'General'
            })
        
        # Sort transactions by date
        transactions.sort(key=lambda x: x['date'], reverse=True)
        
        # Paginate transactions
        paginator = Paginator(transactions, 20)
        transactions_page = paginator.get_page(page)
        
        # Calculate financial metrics
        profit = net_revenue - expenses
        profit_margin = (profit / max(net_revenue, 1)) * 100
        
        return {
            'items': transactions_page,
            'summary': {
                'gross_revenue': revenue,  # Revenue before refunds for orders created in period
                'refunds_amount': total_refunds,  # Use the calculated total refunds
                'net_revenue': net_revenue,  # Revenue after refunds for orders created in period
                'total_expenses': expenses,  # Only approved expenses
                'net_profit': profit,
                'profit_margin': profit_margin,
            },
            'pagination': self._get_pagination_data(transactions_page, paginator)
        }
    
    def _get_daily_summary_data(self, start_date, end_date, page):
        """Get daily summary data based on service orders created each day"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        # Get daily aggregated service order data
        daily_order_data = ServiceOrder.objects.filter(
            created_at__range=[start_datetime, end_datetime]
        ).values('created_at__date').annotate(
            orders_count=Count('id'),
            total_order_value=Sum('total_amount'),
            unique_customers=Count('customer', distinct=True),
            # Get revenue from payments for orders created on this day (regardless of when payment was made)
            revenue_from_payments=Sum(
                'payments__amount',
                filter=Q(
                    payments__status__in=['completed', 'verified']
                ) & ~Q(payments__payment_type='refund')
            ),
            paid_orders_count=Count(
                'payments',
                distinct=True,
                filter=Q(
                    payments__status__in=['completed', 'verified']
                ) & ~Q(payments__payment_type='refund')
            )
        ).order_by('-created_at__date')
        
        # Convert to list for pagination
        daily_list = list(daily_order_data)
        
        # Paginate daily data
        paginator = Paginator(daily_list, 20)
        daily_page = paginator.get_page(page)
        
        # Calculate summary
        total_days = len(daily_list)
        avg_daily_orders = sum(day['orders_count'] for day in daily_list) / max(total_days, 1)
        avg_daily_revenue = sum(day['revenue_from_payments'] or 0 for day in daily_list) / max(total_days, 1)
        total_revenue = sum(day['revenue_from_payments'] or 0 for day in daily_list)
        total_orders = sum(day['orders_count'] for day in daily_list)
        total_order_value = sum(day['total_order_value'] or 0 for day in daily_list)
        
        return {
            'items': daily_page,
            'summary': {
                'total_days': total_days,
                'avg_daily_orders': avg_daily_orders,
                'avg_daily_revenue': avg_daily_revenue,
                'total_revenue': total_revenue,
                'total_orders': total_orders,
                'total_order_value': total_order_value,
                'avg_order_value': total_order_value / max(total_orders, 1),
            },
            'pagination': self._get_pagination_data(daily_page, paginator)
        }
    
    def _get_weekly_summary_data(self, start_date, end_date, page):
        """Get weekly summary data based on service orders created in each week"""
        from django.core.paginator import Paginator
        from datetime import datetime, timedelta
        
        try:
            # Convert dates to datetime for proper filtering
            start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
            
            # Get all service orders in the date range
            orders = ServiceOrder.objects.filter(
                created_at__range=[start_datetime, end_datetime]
            ).select_related('customer').order_by('-created_at')
            
            # Group orders by week manually
            weekly_data = {}
            for order in orders:
                # Get the start of the week (Monday)
                order_date = order.created_at.date()
                week_start = order_date - timedelta(days=order_date.weekday())
                week_key = week_start.strftime('%Y-%m-%d')
                
                if week_key not in weekly_data:
                    weekly_data[week_key] = {
                        'week': week_start,
                        'orders_count': 0,
                        'total_order_value': 0,
                        'revenue_from_payments': 0,
                        'unique_customers': set(),
                        'paid_orders': set()
                    }
                
                weekly_data[week_key]['orders_count'] += 1
                weekly_data[week_key]['total_order_value'] += float(order.total_amount or 0)
                if order.customer_id:
                    weekly_data[week_key]['unique_customers'].add(order.customer_id)
                
                # Add payment revenue for this order (regardless of when payment was made)
                order_payments = Payment.objects.filter(
                    service_order=order,
                    status__in=['completed', 'verified']
                ).exclude(payment_type='refund')
                
                for payment in order_payments:
                    weekly_data[week_key]['revenue_from_payments'] += float(payment.amount or 0)
                    weekly_data[week_key]['paid_orders'].add(order.id)
            
            # Convert sets to counts and prepare final data
            for week_data in weekly_data.values():
                week_data['unique_customers'] = len(week_data['unique_customers'])
                week_data['paid_orders_count'] = len(week_data['paid_orders'])
                del week_data['paid_orders']  # Remove the set, keep only the count
            
            # Convert to list and sort by week
            weekly_list = list(weekly_data.values())
            weekly_list.sort(key=lambda x: x['week'], reverse=True)
            
            # Paginate weekly data
            paginator = Paginator(weekly_list, 20)
            weekly_page = paginator.get_page(page)
            
            # Calculate summary
            total_weeks = len(weekly_list)
            total_orders = sum(week['orders_count'] for week in weekly_list)
            total_revenue = sum(week['revenue_from_payments'] for week in weekly_list)
            avg_weekly_orders = total_orders / max(total_weeks, 1)
            avg_weekly_revenue = total_revenue / max(total_weeks, 1)
            
            return {
                'items': weekly_page,
                'summary': {
                    'total_weeks': total_weeks,
                    'avg_weekly_orders': avg_weekly_orders,
                    'avg_weekly_revenue': avg_weekly_revenue,
                    'total_orders': total_orders,
                    'total_revenue': total_revenue,
                },
                'pagination': self._get_pagination_data(weekly_page, paginator)
            }
        except Exception as e:
            return {
                'items': [],
                'summary': {
                    'total_weeks': 0,
                    'avg_weekly_payments': 0,
                    'avg_weekly_revenue': 0,
                    'total_payments': 0,
                    'total_revenue': 0,
                },
                'pagination': {},
                'error': f'Weekly summary error: {str(e)}'
            }
    
    def _get_monthly_summary_data(self, start_date, end_date, page):
        """Get monthly summary data based on service orders created each month"""
        from django.core.paginator import Paginator
        from datetime import datetime
        
        try:
            # Convert dates to datetime for proper filtering
            start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
            
            # Get all service orders in the date range
            orders = ServiceOrder.objects.filter(
                created_at__range=[start_datetime, end_datetime]
            ).select_related('customer').order_by('-created_at')
            
            # Group orders by month manually
            monthly_data = {}
            for order in orders:
                # Get the first day of the month
                order_date = order.created_at.date()
                month_key = order_date.strftime('%Y-%m')
                month_start = order_date.replace(day=1)
                
                if month_key not in monthly_data:
                    monthly_data[month_key] = {
                        'month': month_start,
                        'orders_count': 0,
                        'total_order_value': 0,
                        'revenue_from_payments': 0,
                        'unique_customers': set(),
                        'paid_orders': set()
                    }
                
                monthly_data[month_key]['orders_count'] += 1
                monthly_data[month_key]['total_order_value'] += float(order.total_amount or 0)
                if order.customer_id:
                    monthly_data[month_key]['unique_customers'].add(order.customer_id)
                
                # Add payment revenue for this order (regardless of when payment was made)
                order_payments = Payment.objects.filter(
                    service_order=order,
                    status__in=['completed', 'verified']
                ).exclude(payment_type='refund')
                
                for payment in order_payments:
                    monthly_data[month_key]['revenue_from_payments'] += float(payment.amount or 0)
                    monthly_data[month_key]['paid_orders'].add(order.id)
            
            # Convert sets to counts
            for month_data in monthly_data.values():
                month_data['unique_customers'] = len(month_data['unique_customers'])
                month_data['paid_orders_count'] = len(month_data['paid_orders'])
                del month_data['paid_orders']  # Remove the set, keep only the count
            
            # Convert to list and sort by month
            monthly_list = list(monthly_data.values())
            monthly_list.sort(key=lambda x: x['month'], reverse=True)
            
            # Paginate monthly data
            paginator = Paginator(monthly_list, 20)
            monthly_page = paginator.get_page(page)
            
            # Calculate summary
            total_months = len(monthly_list)
            total_orders = sum(month['orders_count'] for month in monthly_list)
            total_revenue = sum(month['revenue_from_payments'] for month in monthly_list)
            total_order_value = sum(month['total_order_value'] for month in monthly_list)
            avg_monthly_orders = total_orders / max(total_months, 1)
            avg_monthly_revenue = total_revenue / max(total_months, 1)
            
            return {
                'items': monthly_page,
                'summary': {
                    'total_months': total_months,
                    'avg_monthly_orders': avg_monthly_orders,
                    'avg_monthly_revenue': avg_monthly_revenue,
                    'total_orders': total_orders,
                    'total_revenue': total_revenue,
                    'total_order_value': total_order_value,
                },
                'pagination': self._get_pagination_data(monthly_page, paginator)
            }
        except Exception as e:
            return {
                'items': [],
                'summary': {
                    'total_months': 0,
                    'avg_monthly_payments': 0,
                    'avg_monthly_revenue': 0,
                    'total_payments': 0,
                    'total_revenue': 0,
                },
                'pagination': {},
                'error': f'Monthly summary error: {str(e)}'
            }
    
    def _get_sales_analysis_data(self, start_date, end_date, page):
        """Get sales analysis data based on service orders created in the date range"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        # Get sales by service based on orders created in the date range
        sales_data = ServiceOrderItem.objects.filter(
            order__created_at__range=[start_datetime, end_datetime]
        ).values(
            'service__name', 'service__category__name'
        ).annotate(
            quantity_sold=Sum('quantity'),
            # Total order value for this service from orders created in the period
            total_order_value=Sum('total_price'),
            # For revenue, we'll calculate it as the total_price (order value) for this service
            # This represents the revenue potential of the service
            total_revenue=Sum('total_price'),
            orders_count=Count('order', distinct=True),
            # Count orders that have at least one completed payment
            paid_orders_count=Count('order', distinct=True,
                                  filter=Q(
                                      order__payments__status__in=['completed', 'verified']
                                  ) & ~Q(order__payments__payment_type='refund'))
        ).order_by('-total_order_value')
        
        # Convert to list for pagination
        sales_list = list(sales_data)
        
        # Paginate sales data
        paginator = Paginator(sales_list, 20)
        sales_page = paginator.get_page(page)
        
        # Calculate summary
        total_services = len(sales_list)
        total_sales_revenue = sum(item['total_revenue'] or 0 for item in sales_list)
        total_order_value = sum(item['total_order_value'] or 0 for item in sales_list)
        total_items_sold = sum(item['quantity_sold'] or 0 for item in sales_list)
        total_orders = sum(item['orders_count'] or 0 for item in sales_list)
        total_paid_orders = sum(item['paid_orders_count'] or 0 for item in sales_list)
        
        return {
            'items': sales_page,
            'summary': {
                'total_services': total_services,
                'total_sales_revenue': total_sales_revenue,
                'total_order_value': total_order_value,
                'total_items_sold': total_items_sold,
                'total_orders': total_orders,
                'total_paid_orders': total_paid_orders,
                'avg_revenue_per_service': total_sales_revenue / max(total_services, 1),
                'avg_order_value': total_order_value / max(total_orders, 1),
            },
            'pagination': self._get_pagination_data(sales_page, paginator)
        }
    
    def _get_customer_analytics_data(self, start_date, end_date, page):
        """Get customer analytics data"""
        from django.core.paginator import Paginator
        
        try:
            # Convert dates to datetime for proper filtering
            start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
            
            # Get customer behavior data
            customer_data = Customer.objects.annotate(
                total_orders=Count('service_orders', 
                                 filter=Q(service_orders__created_at__range=[start_datetime, end_datetime])),
                total_spent=Sum('service_orders__total_amount',
                              filter=Q(service_orders__created_at__range=[start_datetime, end_datetime])),
                avg_order_value=Avg('service_orders__total_amount',
                                  filter=Q(service_orders__created_at__range=[start_datetime, end_datetime])),
                last_order_date=Max('service_orders__created_at',
                                  filter=Q(service_orders__created_at__range=[start_datetime, end_datetime]))
            ).filter(total_orders__gt=0).order_by('-total_spent')
            
            # Paginate customer data
            paginator = Paginator(customer_data, 20)
            customer_page = paginator.get_page(page)
            
            # Calculate summary
            total_active_customers = customer_data.count()
            total_customer_value = sum(c.total_spent or 0 for c in customer_data)
            avg_customer_value = total_customer_value / max(total_active_customers, 1)
            
            # Debug information
            debug_info = {
                'customers_found': total_active_customers,
                'total_value': total_customer_value,
                'avg_value': avg_customer_value,
                'date_range': f"{start_date} to {end_date}",
                'query_filter': str(customer_data.query)
            }
            
            return {
                'items': customer_page,
                'summary': {
                    'total_active_customers': total_active_customers,
                    'total_customer_value': total_customer_value,
                    'avg_customer_value': avg_customer_value,
                },
                'pagination': self._get_pagination_data(customer_page, paginator),
                'debug_info': debug_info
            }
        except Exception as e:
            return {
                'items': [],
                'summary': {
                    'total_active_customers': 0,
                    'total_customer_value': 0,
                    'avg_customer_value': 0,
                },
                'pagination': {},
                'error': f"Error generating customer analytics: {str(e)}",
                'debug_info': {'error': str(e)}
            }
    
    def _get_supplier_report_data(self, start_date, end_date, page):
        """Get supplier report data based on approved expenses"""
        from django.core.paginator import Paginator
        
        # Get vendor data from approved expenses in the period
        try:
            vendors = Vendor.objects.annotate(
                approved_expenses_count=Count('expenses',
                                            filter=Q(expenses__expense_date__range=[start_date, end_date],
                                                    expenses__status='approved')),
                total_approved_expenses=Sum('expenses__total_amount',
                                          filter=Q(expenses__expense_date__range=[start_date, end_date],
                                                  expenses__status='approved'))
            ).filter(approved_expenses_count__gt=0).order_by('-total_approved_expenses')
            
            # Paginate vendors
            paginator = Paginator(vendors, 20)
            vendors_page = paginator.get_page(page)
            
            # Calculate summary
            total_vendors = vendors.count()
            total_approved_expenses = sum(v.total_approved_expenses or 0 for v in vendors)
            avg_expense_per_vendor = total_approved_expenses / max(total_vendors, 1)
            
            return {
                'items': vendors_page,
                'summary': {
                    'total_vendors': total_vendors,
                    'total_approved_expenses': total_approved_expenses,
                    'avg_expense_per_vendor': avg_expense_per_vendor,
                },
                'pagination': self._get_pagination_data(vendors_page, paginator)
            }
        except Exception as e:
            # Return empty data if there are issues
            return {
                'items': [],
                'summary': {
                    'total_vendors': 0,
                    'total_approved_expenses': 0,
                    'avg_expense_per_vendor': 0,
                },
                'pagination': {},
                'error': f'Vendor data error: {str(e)}'
            }
    
    # ============== NEW REPORT TYPES ==============
    
    def _get_vehicle_analysis_data(self, start_date, end_date, page):
        """Get vehicle analysis data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # Get vehicles with service frequency based on orders created in the period
            vehicles = Vehicle.objects.select_related('customer').annotate(
                orders_count=Count('service_orders',
                                 filter=Q(service_orders__created_at__range=[start_datetime, end_datetime])),
                total_order_value=Sum('service_orders__total_amount',
                                    filter=Q(service_orders__created_at__range=[start_datetime, end_datetime])),
                revenue_from_payments=Sum('service_orders__payments__amount',
                                        filter=Q(service_orders__created_at__range=[start_datetime, end_datetime],
                                                service_orders__payments__status__in=['completed', 'verified']) &
                                                ~Q(service_orders__payments__payment_type='refund')),
                paid_orders_count=Count('service_orders__payments',
                                      distinct=True,
                                      filter=Q(service_orders__created_at__range=[start_datetime, end_datetime],
                                              service_orders__payments__status__in=['completed', 'verified']) &
                                              ~Q(service_orders__payments__payment_type='refund')),
                last_service_order=Max('service_orders__created_at',
                                     filter=Q(service_orders__created_at__range=[start_datetime, end_datetime]))
            ).filter(orders_count__gt=0).order_by('-orders_count')
            
            # Paginate vehicles
            paginator = Paginator(vehicles, 20)
            vehicles_page = paginator.get_page(page)
            
            # Calculate summary
            total_vehicles = vehicles.count()
            total_orders = sum(v.orders_count or 0 for v in vehicles)
            total_order_value = sum(v.total_order_value or 0 for v in vehicles)
            total_revenue = sum(v.revenue_from_payments or 0 for v in vehicles)
            avg_orders_per_vehicle = total_orders / max(total_vehicles, 1)
            
            return {
                'items': vehicles_page,
                'summary': {
                    'total_vehicles': total_vehicles,
                    'total_orders': total_orders,
                    'total_order_value': total_order_value,
                    'total_revenue': total_revenue,
                    'avg_orders_per_vehicle': avg_orders_per_vehicle,
                },
                'pagination': self._get_pagination_data(vehicles_page, paginator)
            }
        except Exception as e:
            return {'error': f'Vehicle data not available: {str(e)}', 'items': [], 'summary': {}}
    
    def _get_service_packages_data(self, start_date, end_date, page):
        """Get service packages data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # Get service packages with usage statistics based on orders created in period
            packages = ServicePackage.objects.annotate(
                orders_count=Count('serviceorder',
                                 filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime])),
                total_order_value=Sum('serviceorder__total_amount',
                                    filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime])),
                revenue_from_payments=Sum('serviceorder__payments__amount',
                                        filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime],
                                                serviceorder__payments__status__in=['completed', 'verified']) &
                                                ~Q(serviceorder__payments__payment_type='refund')),
                paid_orders_count=Count('serviceorder__payments',
                                      distinct=True,
                                      filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime],
                                              serviceorder__payments__status__in=['completed', 'verified']) &
                                              ~Q(serviceorder__payments__payment_type='refund'))
            ).filter(orders_count__gt=0).order_by('-orders_count')
            
            # Paginate packages
            paginator = Paginator(packages, 20)
            packages_page = paginator.get_page(page)
            
            # Calculate summary
            total_packages = packages.count()
            total_orders = sum(p.orders_count or 0 for p in packages)
            total_order_value = sum(p.total_order_value or 0 for p in packages)
            total_revenue = sum(p.revenue_from_payments or 0 for p in packages)
            
            return {
                'items': packages_page,
                'summary': {
                    'total_packages': total_packages,
                    'total_orders': total_orders,
                    'total_order_value': total_order_value,
                    'total_revenue': total_revenue,
                    'avg_orders_per_package': total_orders / max(total_packages, 1),
                },
                'pagination': self._get_pagination_data(packages_page, paginator)
            }
        except:
            return {'error': 'Service packages data not available', 'items': [], 'summary': {}}
    
    def _get_loyalty_program_data(self, start_date, end_date, page):
        """Get loyalty program data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # Get customers with loyalty activity based on orders created in period
            customers_with_loyalty = Customer.objects.annotate(
                orders_count=Count('serviceorder',
                                 filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime])),
                total_spent=Sum('serviceorder__total_amount',
                              filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime])),
                payments_made=Sum('serviceorder__payments__amount',
                                filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime],
                                        serviceorder__payments__status__in=['completed', 'verified']) &
                                        ~Q(serviceorder__payments__payment_type='refund')),
                points_earned=Sum('serviceorder__total_amount',
                                filter=Q(serviceorder__created_at__range=[start_datetime, end_datetime])) / 10
            ).filter(orders_count__gt=0, loyalty_points__gt=0).order_by('-loyalty_points')
            
            # Paginate customers
            paginator = Paginator(customers_with_loyalty, 20)
            loyalty_page = paginator.get_page(page)
            
            # Calculate loyalty metrics
            total_members = customers_with_loyalty.count()
            total_points = sum(c.loyalty_points or 0 for c in customers_with_loyalty)
            total_orders = sum(c.orders_count or 0 for c in customers_with_loyalty)
            total_revenue = sum(c.payments_made or 0 for c in customers_with_loyalty)
            avg_points_per_member = total_points / max(total_members, 1)
            
            return {
                'items': loyalty_page,
                'summary': {
                    'total_members': total_members,
                    'total_orders': total_orders,
                    'total_points': total_points,
                    'total_revenue': total_revenue,
                    'avg_points_per_member': avg_points_per_member,
                },
                'pagination': self._get_pagination_data(loyalty_page, paginator)
            }
        except Exception as e:
            return {'error': f'Loyalty program data not available: {str(e)}', 'items': [], 'summary': {}}
    
    def _get_attendance_report_data(self, start_date, end_date, page):
        """Get employee attendance data"""
        from django.core.paginator import Paginator
        
        try:
            # Get attendance records
            attendance = Attendance.objects.select_related('employee').filter(
                date__range=[start_date, end_date]
            ).order_by('-date', 'employee__employee_id')
            
            # Paginate attendance
            paginator = Paginator(attendance, 20)
            attendance_page = paginator.get_page(page)
            
            # Calculate summary
            total_records = attendance.count()
            present_count = attendance.filter(status='present').count()
            absent_count = attendance.filter(status='absent').count()
            late_count = attendance.filter(status='late').count()
            attendance_rate = (present_count / max(total_records, 1)) * 100
            
            return {
                'items': attendance_page,
                'summary': {
                    'total_records': total_records,
                    'present_count': present_count,
                    'absent_count': absent_count,
                    'late_count': late_count,
                    'attendance_rate': attendance_rate,
                },
                'pagination': self._get_pagination_data(attendance_page, paginator)
            }
        except:
            return {'error': 'Attendance data not available', 'items': [], 'summary': {}}
    
    def _get_department_performance_data(self, start_date, end_date, page):
        """Get department performance data based on orders created in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # Get departments with performance metrics based on orders created in period
            departments = Department.objects.annotate(
                employee_count=Count('employees', filter=Q(employees__is_active=True)),
                orders_handled=Count('employees__assigned_orders',
                                   filter=Q(employees__assigned_orders__created_at__range=[start_datetime, end_datetime])),
                total_order_value=Sum('employees__assigned_orders__total_amount',
                                    filter=Q(employees__assigned_orders__created_at__range=[start_datetime, end_datetime])),
                revenue_from_payments=Sum('employees__assigned_orders__payments__amount',
                                        filter=Q(employees__assigned_orders__created_at__range=[start_datetime, end_datetime],
                                                employees__assigned_orders__payments__status__in=['completed', 'verified']) &
                                                ~Q(employees__assigned_orders__payments__payment_type='refund')),
                avg_order_value=Avg('employees__assigned_orders__total_amount',
                                  filter=Q(employees__assigned_orders__created_at__range=[start_datetime, end_datetime]))
            ).filter(orders_handled__gt=0).order_by('-revenue_from_payments')
            
            # Paginate departments
            paginator = Paginator(departments, 20)
            departments_page = paginator.get_page(page)
            
            # Calculate summary
            total_departments = departments.count()
            total_employees = sum(d.employee_count or 0 for d in departments)
            total_orders = sum(d.orders_handled or 0 for d in departments)
            total_order_value = sum(d.total_order_value or 0 for d in departments)
            total_revenue = sum(d.revenue_from_payments or 0 for d in departments)
            
            return {
                'items': departments_page,
                'summary': {
                    'total_departments': total_departments,
                    'total_employees': total_employees,
                    'total_orders': total_orders,
                    'total_order_value': total_order_value,
                    'total_revenue': total_revenue,
                    'avg_revenue_per_dept': total_revenue / max(total_departments, 1),
                },
                'pagination': self._get_pagination_data(departments_page, paginator)
            }
        except Exception as e:
            return {'error': f'Department data not available: {str(e)}', 'items': [], 'summary': {}}
    
    def _get_subscription_analysis_data(self, start_date, end_date, page):
        """Get subscription analysis data based on subscriptions created and order activity in the period"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # Get subscriptions with activity based on orders created in period
            subscriptions = Subscription.objects.select_related('customer').annotate(
                orders_count=Count('customer__serviceorder',
                                 filter=Q(customer__serviceorder__created_at__range=[start_datetime, end_datetime])),
                total_order_value=Sum('customer__serviceorder__total_amount',
                                    filter=Q(customer__serviceorder__created_at__range=[start_datetime, end_datetime])),
                revenue_from_payments=Sum('customer__serviceorder__payments__amount',
                                        filter=Q(customer__serviceorder__created_at__range=[start_datetime, end_datetime],
                                                customer__serviceorder__payments__status__in=['completed', 'verified']) &
                                                ~Q(customer__serviceorder__payments__payment_type='refund'))
            ).filter(
                Q(created_at__range=[start_datetime, end_datetime]) |
                Q(orders_count__gt=0)
            ).order_by('-revenue_from_payments')
            
            # Paginate subscriptions
            paginator = Paginator(subscriptions, 20)
            subscriptions_page = paginator.get_page(page)
            
            # Calculate summary
            total_subscriptions = subscriptions.count()
            new_subscriptions = subscriptions.filter(created_at__range=[start_datetime, end_datetime]).count()
            active_subscriptions = subscriptions.filter(status='active').count()
            total_subscription_revenue = subscriptions.aggregate(Sum('amount'))['amount__sum'] or 0
            total_order_revenue = sum(s.revenue_from_payments or 0 for s in subscriptions)
            avg_orders_per_subscriber = sum(s.orders_count or 0 for s in subscriptions) / max(total_subscriptions, 1)
            
            return {
                'items': subscriptions_page,
                'summary': {
                    'total_subscriptions': total_subscriptions,
                    'new_subscriptions': new_subscriptions,
                    'active_subscriptions': active_subscriptions,
                    'total_subscription_revenue': total_subscription_revenue,
                    'total_order_revenue': total_order_revenue,
                    'avg_orders_per_subscriber': avg_orders_per_subscriber,
                },
                'pagination': self._get_pagination_data(subscriptions_page, paginator)
            }
        except Exception as e:
            return {'error': f'Subscription data not available: {str(e)}', 'items': [], 'summary': {}}
    
    def _get_balance_sheet_data(self, start_date, end_date, page):
        """Get comprehensive balance sheet data - Assets, Liabilities, Equity"""
        from django.core.paginator import Paginator
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # ASSETS SECTION
            
            # Current Assets - Cash and Cash Equivalents
            cash_from_payments = Payment.objects.filter(
                service_order__created_at__range=[start_datetime, end_datetime],
                status__in=['completed', 'verified'],
                payment_method__name__icontains='cash'
            ).exclude(payment_type='refund').aggregate(
                total=Sum('amount')
            )['total'] or 0
            
            bank_from_payments = Payment.objects.filter(
                service_order__created_at__range=[start_datetime, end_datetime],
                status__in=['completed', 'verified']
            ).exclude(
                payment_method__name__icontains='cash'
            ).exclude(payment_type='refund').aggregate(
                total=Sum('amount')
            )['total'] or 0
            
            # Accounts Receivable - Unpaid service orders
            accounts_receivable = ServiceOrder.objects.filter(
                created_at__range=[start_datetime, end_datetime]
            ).aggregate(
                total_orders=Sum('total_amount'),
                paid_amount=Sum('payments__amount', filter=Q(
                    payments__status__in=['completed', 'verified']
                ) & ~Q(payments__payment_type='refund'))
            )
            
            total_orders_value = Decimal(str(accounts_receivable['total_orders'] or 0))
            total_paid = Decimal(str(accounts_receivable['paid_amount'] or 0))
            accounts_receivable_amount = total_orders_value - total_paid
            
            # Inventory Assets
            inventory_value = Decimal('0')
            try:
                from apps.inventory.models import InventoryItem
                inventory_items = InventoryItem.objects.all()
                for item in inventory_items:
                    unit_cost = getattr(item, 'unit_cost', 0) or 0
                    current_stock = getattr(item, 'current_stock', 0) or 0
                    # Convert to Decimal to avoid type mixing
                    inventory_value += Decimal(str(unit_cost)) * Decimal(str(current_stock))
            except Exception:
                inventory_value = Decimal('0')
                
            # Total Current Assets
            total_current_assets = cash_from_payments + bank_from_payments + accounts_receivable_amount + inventory_value
            
            # LIABILITIES SECTION
            
            # Current Liabilities - Accounts Payable (Unpaid Approved Expenses)
            accounts_payable = Expense.objects.filter(
                created_at__range=[start_datetime, end_datetime],
                status='approved'
            ).aggregate(
                total_approved=Sum('amount'),
                total_paid=Sum('amount', filter=Q(status='paid'))
            )
            
            total_approved_expenses = Decimal(str(accounts_payable['total_approved'] or 0))
            total_paid_expenses = Decimal(str(accounts_payable['total_paid'] or 0))
            accounts_payable_amount = total_approved_expenses - total_paid_expenses
            
            # Employee Liabilities (Unpaid wages/salaries)
            employee_liabilities = Decimal('0')
            try:
                # Calculate based on employee performance and unpaid wages
                employees_with_revenue = Employee.objects.filter(is_active=True).annotate(
                    revenue_generated=Sum('assigned_orders__payments__amount',
                                        filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime],
                                                assigned_orders__payments__status__in=['completed', 'verified']) &
                                                ~Q(assigned_orders__payments__payment_type='refund'))
                )
                # Estimate 10% commission as liability
                for emp in employees_with_revenue:
                    if emp.revenue_generated:
                        # Convert to Decimal to avoid type mixing
                        employee_liabilities += Decimal(str(emp.revenue_generated)) * Decimal('0.1')
            except Exception:
                employee_liabilities = Decimal('0')
            
            # Total Current Liabilities
            total_current_liabilities = accounts_payable_amount + employee_liabilities
            
            # EQUITY SECTION
            
            # Revenue
            total_revenue = cash_from_payments + bank_from_payments
            
            # Total Expenses (All paid expenses in period)
            total_expenses = Expense.objects.filter(
                created_at__range=[start_datetime, end_datetime],
                status__in=['approved', 'paid']
            ).aggregate(total=Sum('amount'))['total'] or 0
            total_expenses = Decimal(str(total_expenses))
            
            # Net Income
            net_income = total_revenue - total_expenses
            
            # Owner's Equity (simplified)
            owners_equity = net_income
            
            # Total Equity
            total_equity = owners_equity
            
            # BALANCE SHEET ITEMS (for display)
            balance_sheet_items = [
                # Assets
                {'category': 'CURRENT ASSETS', 'name': 'Cash', 'amount': cash_from_payments, 'type': 'asset'},
                {'category': 'CURRENT ASSETS', 'name': 'Bank/Digital Payments', 'amount': bank_from_payments, 'type': 'asset'},
                {'category': 'CURRENT ASSETS', 'name': 'Accounts Receivable', 'amount': accounts_receivable_amount, 'type': 'asset'},
                {'category': 'CURRENT ASSETS', 'name': 'Inventory', 'amount': inventory_value, 'type': 'asset'},
                {'category': 'TOTAL ASSETS', 'name': 'Total Current Assets', 'amount': total_current_assets, 'type': 'asset_total'},
                
                # Liabilities
                {'category': 'CURRENT LIABILITIES', 'name': 'Accounts Payable', 'amount': accounts_payable_amount, 'type': 'liability'},
                {'category': 'CURRENT LIABILITIES', 'name': 'Employee Liabilities', 'amount': employee_liabilities, 'type': 'liability'},
                {'category': 'TOTAL LIABILITIES', 'name': 'Total Current Liabilities', 'amount': total_current_liabilities, 'type': 'liability_total'},
                
                # Equity
                {'category': 'EQUITY', 'name': 'Net Income', 'amount': net_income, 'type': 'equity'},
                {'category': 'EQUITY', 'name': "Owner's Equity", 'amount': owners_equity, 'type': 'equity'},
                {'category': 'TOTAL EQUITY', 'name': 'Total Equity', 'amount': total_equity, 'type': 'equity_total'},
            ]
            
            # Paginate balance sheet items
            paginator = Paginator(balance_sheet_items, 20)
            balance_sheet_page = paginator.get_page(page)
            
            # Verify Balance Sheet Equation: Assets = Liabilities + Equity
            balance_check = total_current_assets - (total_current_liabilities + total_equity)
            
            return {
                'items': balance_sheet_page,
                'summary': {
                    'total_assets': total_current_assets,
                    'total_liabilities': total_current_liabilities,
                    'total_equity': total_equity,
                    'total_revenue': total_revenue,
                    'total_expenses': total_expenses,
                    'net_income': net_income,
                    'balance_check': balance_check,
                    'balance_verified': abs(balance_check) < 0.01,  # Within 1 cent
                    'cash_on_hand': cash_from_payments,
                    'bank_balance': bank_from_payments,
                    'outstanding_receivables': accounts_receivable_amount,
                    'inventory_value': inventory_value,
                    'outstanding_payables': accounts_payable_amount,
                },
                'pagination': self._get_pagination_data(balance_sheet_page, paginator)
            }
            
        except Exception as e:
            return {'error': f'Balance sheet data not available: {str(e)}', 'items': [], 'summary': {}}

    def _get_refunds_report_data(self, start_date, end_date, page):
        """Get comprehensive refunds and returns data"""
        from django.core.paginator import Paginator
        from apps.payments.models import PaymentRefund
        
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        try:
            # Get all refund-related data
            refund_items = []
            processed_orders = set()  # Track orders already processed to avoid duplicates
            
            # 1. Payment records with payment_type='refund' for orders created in period
            refund_payments = Payment.objects.select_related(
                'service_order', 'customer', 'payment_method', 'processed_by'
            ).filter(
                service_order__created_at__range=[start_datetime, end_datetime],
                payment_type='refund',
                status__in=['completed', 'verified']
            ).order_by('-completed_at')
            
            for payment in refund_payments:
                order_key = f"{payment.service_order.id}_{payment.id}" if payment.service_order else f"no_order_{payment.id}"
                if order_key not in processed_orders:
                    processed_orders.add(order_key)
                    refund_items.append({
                        'type': 'Direct Refund Payment',
                        'refund_id': payment.payment_id,
                        'original_order': payment.service_order.order_number if payment.service_order else 'N/A',
                        'customer': payment.customer.full_name if payment.customer else 'N/A',
                        'amount': payment.amount,
                        'reason': 'Payment refund',
                        'processed_by': payment.processed_by.full_name if payment.processed_by else 'System',
                        'processed_at': payment.completed_at,
                        'status': 'Completed',
                        'payment_method': payment.payment_method.name if payment.payment_method else 'N/A',
                        'order_date': payment.service_order.created_at if payment.service_order else None,
                    })
            
            # 2. PaymentRefund records for orders created in period
            payment_refunds = PaymentRefund.objects.select_related(
                'original_payment', 'original_payment__service_order', 
                'original_payment__customer', 'processed_by', 'approved_by'
            ).filter(
                original_payment__service_order__created_at__range=[start_datetime, end_datetime],
                status='completed'
            ).order_by('-processed_at')
            
            for refund in payment_refunds:
                # Check if this refund's original payment was already processed
                original_order_key = f"{refund.original_payment.service_order.id}_{refund.original_payment.id}" if refund.original_payment.service_order else f"no_order_{refund.original_payment.id}"
                if original_order_key not in processed_orders:
                    processed_orders.add(original_order_key)
                    refund_items.append({
                        'type': 'Formal Refund Record',
                        'refund_id': refund.refund_id,
                        'original_order': refund.original_payment.service_order.order_number if refund.original_payment.service_order else 'N/A',
                        'customer': refund.original_payment.customer.full_name if refund.original_payment.customer else 'N/A',
                        'amount': refund.amount,
                        'reason': refund.reason[:100] + '...' if len(refund.reason) > 100 else refund.reason,
                        'processed_by': refund.processed_by.full_name if refund.processed_by else 'N/A',
                        'processed_at': refund.processed_at,
                        'status': refund.get_status_display(),
                        'payment_method': refund.original_payment.payment_method.name if refund.original_payment.payment_method else 'N/A',
                        'order_date': refund.original_payment.service_order.created_at if refund.original_payment.service_order else None,
                    })
            
            # 3. Payments with status='refunded' for orders created in period (only if not already processed)
            refunded_payments = Payment.objects.select_related(
                'service_order', 'customer', 'payment_method'
            ).filter(
                service_order__created_at__range=[start_datetime, end_datetime],
                status='refunded'
            ).order_by('-updated_at')
            
            for payment in refunded_payments:
                order_key = f"{payment.service_order.id}_{payment.id}" if payment.service_order else f"no_order_{payment.id}"
                if order_key not in processed_orders:
                    processed_orders.add(order_key)
                    refund_items.append({
                        'type': 'Refunded Payment',
                        'refund_id': f"RF-{payment.payment_id}",
                        'original_order': payment.service_order.order_number if payment.service_order else 'N/A',
                        'customer': payment.customer.full_name if payment.customer else 'N/A',
                        'amount': payment.amount,
                        'reason': 'Payment marked as refunded',
                        'processed_by': 'System',
                        'processed_at': payment.updated_at,
                        'status': 'Refunded',
                        'payment_method': payment.payment_method.name if payment.payment_method else 'N/A',
                        'order_date': payment.service_order.created_at if payment.service_order else None,
                    })
            
            # Sort all refunds by processed date
            refund_items.sort(key=lambda x: x['processed_at'] or x['order_date'], reverse=True)
            
            # Paginate refund items
            paginator = Paginator(refund_items, 20)
            refunds_page = paginator.get_page(page)
            
            # Calculate summary
            total_refunds_count = len(refund_items)
            total_refund_amount = sum(item['amount'] for item in refund_items)
            
            # Group by refund type
            refund_types = {}
            for item in refund_items:
                refund_type = item['type']
                if refund_type not in refund_types:
                    refund_types[refund_type] = {'count': 0, 'amount': 0}
                refund_types[refund_type]['count'] += 1
                refund_types[refund_type]['amount'] += item['amount']
            
            # Get revenue for comparison
            orders_in_period = ServiceOrder.objects.filter(
                created_at__range=[start_datetime, end_datetime]
            )
            total_revenue = Payment.objects.filter(
                service_order__in=orders_in_period,
                status__in=['completed', 'verified']
            ).exclude(payment_type='refund').aggregate(Sum('amount'))['amount__sum'] or 0
            
            refund_rate = (total_refund_amount / max(total_revenue, 1)) * 100
            
            return {
                'items': refunds_page,
                'summary': {
                    'total_refunds_count': total_refunds_count,
                    'total_refund_amount': total_refund_amount,
                    'total_revenue': total_revenue,
                    'refund_rate': refund_rate,
                    'refund_types': refund_types,
                    'avg_refund_amount': total_refund_amount / max(total_refunds_count, 1),
                },
                'pagination': self._get_pagination_data(refunds_page, paginator)
            }
            
        except Exception as e:
            return {'error': f'Refunds data not available: {str(e)}', 'items': [], 'summary': {}}

    # ============== HELPER METHODS ==============
    
    def _get_pagination_data(self, page_obj, paginator):
        """Get pagination data for templates"""
        return {
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
            'num_pages': paginator.num_pages,
            'page_range': paginator.page_range,
        }
    
    # ============== EXPORT HANDLING ==============
    
    def _handle_export(self, request, export_format):
        """Handle PDF and Excel exports with focused report data"""
        report_type = request.GET.get('report_type', 'business_overview')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        employee_id = request.GET.get('employee_id')  # Add employee_id parameter
        
        # Set default date range if not provided
        if not start_date or not end_date:
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get all report data (not paginated for export)
        report_data = self._get_export_data(report_type, start_date, end_date, employee_id)
        
        if export_format == 'pdf':
            return self._generate_pdf_export(report_data, request.tenant)
        else:  # excel
            return self._generate_excel_export(report_data, request.tenant)
    
    def _get_export_data(self, report_type, start_date, end_date, employee_id=None):
        """Get all data for export (no pagination) - Updated to use order-based data"""
        # Convert dates to datetime for proper filtering
        start_datetime, end_datetime = self._get_datetime_range(start_date, end_date)
        
        data = {
            'title': self._get_report_title(report_type),
            'type': report_type,
            'period': f"{start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}",
            'items': [],
            'summary': {},
            'employee_id': employee_id  # Pass employee_id to data
        }
        
        try:
            # Get data without pagination by calling the same methods with page=1 and extracting items
            if report_type == 'individual_employee' and employee_id:
                paginated_data = self._get_individual_employee_data(employee_id, start_date, end_date, 1)
            else:
                paginated_data = self._get_report_data(report_type, start_date, end_date, 1)
            
            # For export, we want all items, not just the paginated ones - Updated to use order-based data
            if report_type == 'business_overview':
                # Use service orders created in the period
                data['items'] = list(ServiceOrder.objects.select_related(
                    'customer', 
                    'assigned_attendant', 
                    'vehicle'
                ).filter(
                    created_at__range=[start_datetime, end_datetime]
                ).order_by('-created_at'))
            elif report_type == 'inventory':
                # Get inventory items with movement data
                items = InventoryItem.objects.select_related('category', 'unit').annotate(
                    total_in=Sum('stock_movements__quantity', 
                                filter=Q(stock_movements__created_at__range=[start_datetime, end_datetime],
                                       stock_movements__movement_type='in')),
                    total_out=Sum('stock_movements__quantity', 
                                 filter=Q(stock_movements__created_at__range=[start_datetime, end_datetime],
                                        stock_movements__movement_type='out')),
                    movement_count=Count('stock_movements', 
                                       filter=Q(stock_movements__created_at__range=[start_datetime, end_datetime]))
                ).order_by('-movement_count', 'name')
                data['items'] = list(items)
            elif report_type == 'services':
                # Get services with performance metrics based on orders created in period
                services = Service.objects.select_related('category').annotate(
                    orders_count=Count('order_items__order',
                                     filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime])),
                    total_revenue=Sum('order_items__order__payments__amount',
                                    filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime],
                                            order_items__order__payments__status__in=['completed', 'verified']) &
                                            ~Q(order_items__order__payments__payment_type='refund')),
                    avg_rating=Avg('order_items__rating',
                                 filter=Q(order_items__order__created_at__range=[start_datetime, end_datetime]))
                ).order_by('-total_revenue', 'name')
                data['items'] = list(services)
            elif report_type == 'customers':
                # Get customers with their order activity in the date range
                customers = Customer.objects.annotate(
                    total_orders=Count('service_orders', 
                                     filter=Q(service_orders__created_at__range=[start_datetime, end_datetime])),
                    total_spent=Sum('service_orders__payments__amount',
                                  filter=Q(service_orders__created_at__range=[start_datetime, end_datetime],
                                          service_orders__payments__status__in=['completed', 'verified']) &
                                          ~Q(service_orders__payments__payment_type='refund')),
                    avg_order_value=Avg('service_orders__total_amount',
                                      filter=Q(service_orders__created_at__range=[start_datetime, end_datetime]))
                ).filter(total_orders__gt=0).order_by('-total_spent')
                data['items'] = list(customers)
            elif report_type == 'employees':
                # Get employees with their performance metrics based on orders created in period
                employees = Employee.objects.select_related('department', 'position').annotate(
                    orders_assigned_count=Count('assigned_orders',
                                              filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime])),
                    total_revenue_generated=Sum('assigned_orders__payments__amount',
                                              filter=Q(assigned_orders__created_at__range=[start_datetime, end_datetime],
                                                      assigned_orders__payments__status__in=['completed', 'verified']) &
                                                      ~Q(assigned_orders__payments__payment_type='refund'))
                ).filter(is_active=True).order_by('-total_revenue_generated')
                data['items'] = list(employees)
            elif report_type == 'payments':
                # Get payments for orders created in the period
                data['items'] = list(Payment.objects.select_related('payment_method', 'service_order', 'customer').filter(
                    service_order__created_at__range=[start_datetime, end_datetime],
                    status__in=['completed', 'verified']
                ).exclude(
                    payment_type='refund'
                ).order_by('-completed_at', '-created_at'))
            elif report_type == 'expenses':
                # Get only approved expenses
                data['items'] = list(Expense.objects.select_related('category', 'vendor').filter(
                    expense_date__range=[start_date, end_date],
                    status='approved'
                ).order_by('-expense_date'))
            elif report_type == 'individual_employee':
                # For individual employee, we need to include employee_details and payment data
                if employee_id:
                    try:
                        employee = Employee.objects.get(id=employee_id)
                        data['employee_details'] = {
                            'employee_id': employee.employee_id,
                            'full_name': employee.full_name,
                            'email': employee.email or 'N/A',
                            'phone_number': str(getattr(employee, 'phone', 'N/A')),
                            'department': str(employee.department) if employee.department else 'N/A',
                            'position': str(employee.position) if employee.position else 'N/A',
                            'hire_date': employee.hire_date,
                            'is_active': employee.is_active,
                            'role': employee.get_role_display() if employee.role else 'N/A'
                        }
                        data['employee'] = employee
                        
                        # Get payment data for this employee
                        payments = Payment.objects.select_related(
                            'payment_method', 'service_order', 'service_order__customer'
                        ).filter(
                            service_order__assigned_attendant=employee,
                            completed_at__range=[start_datetime, end_datetime],
                            status__in=['completed', 'verified']
                        ).exclude(
                            payment_type='refund'
                        ).order_by('-completed_at')
                        data['items'] = list(payments)
                        
                        # Add payment methods breakdown
                        payment_methods = Payment.objects.filter(
                            service_order__assigned_attendant=employee,
                            completed_at__range=[start_datetime, end_datetime],
                            status__in=['completed', 'verified']
                        ).exclude(
                            payment_type='refund'
                        ).values('payment_method__name').annotate(
                            total_amount=Sum('amount')
                        ).order_by('-total_amount')
                        data['payment_methods'] = list(payment_methods)
                        
                        # Get summary data from paginated data
                        data['summary'] = paginated_data.get('summary', {})
                        
                    except Employee.DoesNotExist:
                        data['error'] = "Employee not found"
                else:
                    data['error'] = "Employee ID not provided"
            elif report_type == 'expenses':
                data['items'] = list(Expense.objects.select_related('category', 'vendor').filter(
                    expense_date__range=[start_date, end_date]
                ).order_by('-expense_date'))
            elif report_type == 'vehicle_analysis':
                # Get vehicles with payment-based service data
                vehicles = Vehicle.objects.select_related('customer').annotate(
                    paid_services_count=Count('service_orders__payments',
                                            filter=Q(service_orders__payments__completed_at__range=[start_datetime, end_datetime],
                                                    service_orders__payments__status__in=['completed', 'verified']) &
                                                    ~Q(service_orders__payments__payment_type='refund')),
                    total_spent=Sum('service_orders__payments__amount',
                                  filter=Q(service_orders__payments__completed_at__range=[start_datetime, end_datetime],
                                          service_orders__payments__status__in=['completed', 'verified']) &
                                          ~Q(service_orders__payments__payment_type='refund')),
                    last_service=Max('service_orders__payments__completed_at',
                                   filter=Q(service_orders__payments__completed_at__range=[start_datetime, end_datetime],
                                           service_orders__payments__status__in=['completed', 'verified']) &
                                           ~Q(service_orders__payments__payment_type='refund'))
                ).filter(paid_services_count__gt=0).order_by('-paid_services_count')
                data['items'] = list(vehicles)
            elif report_type in ['daily_summary', 'weekly_summary', 'monthly_summary', 'sales_analysis']:
                # For summary reports, use paginated data which is already payment-based
                data['items'] = list(paginated_data.get('items', []))
            else:
                # For other report types, use the paginated data items
                data['items'] = list(paginated_data.get('items', []))
            
            data['summary'] = paginated_data.get('summary', {})
            
        except Exception as e:
            data['error'] = f"Error generating {report_type} export: {str(e)}"
            
        return data
    
    def _generate_pdf_export(self, report_data, tenant):
        """Generate comprehensive PDF for specific report type"""
        if not PDF_AVAILABLE:
            return JsonResponse({'error': 'PDF generation not available'}, status=500)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.8*inch, bottomMargin=1*inch)
        styles = getSampleStyleSheet()
        story = []
        
        # Header with tenant information
        header_data = [
            [tenant.name, f"Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"],
            [getattr(tenant, 'address', 'Address not available'), f"Report Period: {report_data.get('period', '')}"],
            [getattr(tenant, 'phone', ''), '']
        ]
        header_table = Table(header_data, colWidths=[4*inch, 2.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 20))
        
        # Main Title
        title = Paragraph(f"<b>{report_data.get('title', 'Business Report')}</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Add employee details for individual employee report
        if report_data.get('type') == 'individual_employee' and report_data.get('employee_details'):
            employee_details = report_data['employee_details']
            employee_title = Paragraph("<b>Employee Information</b>", styles['Heading2'])
            story.append(employee_title)
            story.append(Spacer(1, 10))
            
            # Create employee details table
            employee_data = [
                ['Employee ID:', employee_details.get('employee_id', 'N/A')],
                ['Full Name:', employee_details.get('full_name', 'N/A')],
                ['Email:', employee_details.get('email', 'N/A')],
                ['Phone:', employee_details.get('phone_number', 'N/A')],
                ['Department:', employee_details.get('department', 'N/A')],
                ['Position:', str(employee_details.get('position', 'N/A'))],
                ['Hire Date:', str(employee_details.get('hire_date', 'N/A'))],
                ['Status:', 'Active' if employee_details.get('is_active') else 'Inactive']
            ]
            
            employee_table = Table(employee_data, colWidths=[2*inch, 3*inch])
            employee_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (1, 0), (1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(employee_table)
            story.append(Spacer(1, 20))
        
        # Summary section
        if report_data.get('summary'):
            summary_title = Paragraph("<b>Summary</b>", styles['Heading2'])
            story.append(summary_title)
            story.append(Spacer(1, 10))
            
            summary_data = []
            for key, value in report_data['summary'].items():
                if key != 'error' and key != 'refund_types' and value is not None:
                    formatted_key = key.replace('_', ' ').title()
                    if 'revenue' in key.lower() or 'profit' in key.lower() or 'value' in key.lower() or 'amount' in key.lower():
                        formatted_value = f"KES {value:,.2f}"
                    elif 'rate' in key.lower() or 'margin' in key.lower():
                        formatted_value = f"{value:.1f}%" if isinstance(value, (int, float)) else str(value)
                    else:
                        formatted_value = str(value)
                    summary_data.append([formatted_key, formatted_value])
            
            if summary_data:
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 30))
        
        # Detailed Data Table
        if report_data.get('items'):
            data_title = Paragraph("<b>Detailed Data</b>", styles['Heading2'])
            story.append(data_title)
            story.append(Spacer(1, 10))
            
            # Generate table data based on report type
            table_data = self._generate_pdf_table_data(report_data)
            
            if table_data:
                # Create table with appropriate column widths
                col_count = len(table_data[0]) if table_data else 0
                col_width = 6.5*inch / col_count if col_count > 0 else 1*inch
                
                data_table = Table(table_data, colWidths=[col_width] * col_count)
                data_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
                ]))
                story.append(data_table)
        
        # Add signature section for individual employee reports
        if report_data.get('type') == 'individual_employee' and report_data.get('employee'):
            story.append(Spacer(1, 30))
            story.append(Paragraph("<b>Signatures</b>", styles['Heading2']))
            story.append(Spacer(1, 20))
            
            # Signature table
            signature_data = [
                ['Employee Name:', report_data['employee'].full_name, 'Supervisor:'],
                ['', '', ''],
                ['Signature: ___________________', '', 'Signature: ___________________'],
                ['', '', ''],
                [f"Date: {timezone.now().strftime('%B %d, %Y')}", '', f"Date: {timezone.now().strftime('%B %d, %Y')}"]
            ]
            signature_table = Table(signature_data, colWidths=[2.5*inch, 1*inch, 2.5*inch])
            signature_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(signature_table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Report generated by {tenant.name} - AutoWash Management System"
        footer = Paragraph(footer_text, styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{tenant.name}_{report_data.get("type", "report")}_report.pdf"'
        return response
    
    def _generate_pdf_table_data(self, report_data):
        """Generate table data for PDF based on report type - Updated for payment-based data"""
        report_type = report_data.get('type')
        items = report_data.get('items', [])
        
        if not items:
            return []
        
        if report_type == 'business_overview':
            headers = ['Order Date', 'Order #', 'Customer', 'Vehicle', 'Total Amount', 'Payment Status']
            table_data = [headers]
            total_amount = 0
            for item in items[:50]:  # Limit to 50 items for PDF
                amount = float(item.total_amount or 0)
                total_amount += amount
                payment_status = 'Paid' if hasattr(item, 'payment_received') and item.payment_received else 'Unpaid'
                table_data.append([
                    item.created_at.strftime('%Y-%m-%d') if item.created_at else 'N/A',
                    item.order_number if hasattr(item, 'order_number') else 'N/A',
                    item.customer.full_name if item.customer else 'N/A',
                    f"{item.vehicle.make} {item.vehicle.model}" if item.vehicle else 'N/A',
                    f"KES {amount:,.2f}",
                    payment_status
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_amount:,.2f}", ''])
                
        elif report_type == 'inventory':
            headers = ['Item Name', 'Category', 'Current Stock', 'Stock Value', 'Movements']
            table_data = [headers]
            total_value = 0
            for item in items[:50]:
                stock_value = float(item.current_stock or 0) * float(getattr(item, 'unit_cost', 0) or 0)
                total_value += stock_value
                table_data.append([
                    item.name,
                    item.category.name if item.category else 'N/A',
                    str(item.current_stock),
                    f"KES {stock_value:,.2f}",
                    str(getattr(item, 'movement_count', 0) or 0)
                ])
            # Add total row
            table_data.append(['', '', 'TOTAL VALUE:', f"KES {total_value:,.2f}", ''])
                
        elif report_type == 'services':
            headers = ['Service Name', 'Category', 'Price', 'Payments', 'Revenue']
            table_data = [headers]
            total_revenue = 0
            for item in items[:50]:
                revenue = float(getattr(item, 'total_revenue', 0) or 0)
                total_revenue += revenue
                table_data.append([
                    item.name,
                    item.category.name if item.category else 'N/A',
                    f"KES {item.base_price:,.2f}",
                    str(getattr(item, 'payments_count', 0) or 0),
                    f"KES {revenue:,.2f}"
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_revenue:,.2f}"])
                
        elif report_type == 'customers' or report_type == 'customer_analytics':
            headers = ['Customer Name', 'Email', 'Phone', 'Payments', 'Total Spent']
            table_data = [headers]
            total_spent = 0
            for item in items[:50]:
                spent = float(getattr(item, 'total_spent', 0) or 0)
                total_spent += spent
                table_data.append([
                    item.full_name,
                    item.email or 'N/A',
                    str(getattr(item, 'phone', 'N/A')) or 'N/A',
                    str(getattr(item, 'total_payments', 0) or 0),
                    f"KES {spent:,.2f}"
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_spent:,.2f}"])
                
        elif report_type == 'employees':
            headers = ['Employee Name', 'Department', 'Position', 'Payments', 'Revenue']
            table_data = [headers]
            total_revenue = 0
            for item in items[:50]:
                revenue = float(getattr(item, 'revenue_generated', 0) or 0)
                total_revenue += revenue
                table_data.append([
                    item.full_name,
                    item.department.name if item.department else 'N/A',
                    item.position.title if item.position else 'N/A',
                    str(getattr(item, 'payments_handled', 0) or 0),
                    f"KES {revenue:,.2f}"
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_revenue:,.2f}"])
                
        elif report_type == 'individual_employee':
            employee = report_data.get('employee')
            if employee:
                headers = ['Payment Date', 'Order #', 'Customer', 'Payment Method', 'Amount', 'Status']
                table_data = [headers]
                total_amount = 0
                for item in items[:50]:
                    amount = float(item.amount or 0)
                    total_amount += amount
                    table_data.append([
                        item.completed_at.strftime('%Y-%m-%d') if item.completed_at else 'N/A',
                        item.service_order.order_number if item.service_order else 'N/A',
                        item.service_order.customer.full_name if item.service_order and item.service_order.customer else 'N/A',
                        item.payment_method.name if item.payment_method else 'Cash',
                        f"KES {amount:,.2f}",
                        item.status.title()
                    ])
                # Add total row
                table_data.append(['', '', '', 'TOTAL:', f"KES {total_amount:,.2f}", ''])
            else:
                return []
                
        elif report_type == 'financial_summary':
            headers = ['Date', 'Type', 'Description', 'Category', 'Amount']
            table_data = [headers]
            total_revenue = 0
            total_expenses = 0
            for item in items[:50]:
                amount = float(item.get('amount', 0) or 0)
                if item.get('type') == 'Revenue':
                    total_revenue += amount
                else:
                    total_expenses += abs(amount)
                table_data.append([
                    item.get('date', '').strftime('%Y-%m-%d') if hasattr(item.get('date', ''), 'strftime') else str(item.get('date', '')),
                    item.get('type', 'N/A'),
                    item.get('description', 'N/A')[:40],
                    item.get('category', 'N/A'),
                    f"KES {amount:,.2f}"
                ])
            # Add summary rows
            table_data.append(['', '', '', 'TOTAL REVENUE:', f"KES {total_revenue:,.2f}"])
            table_data.append(['', '', '', 'TOTAL EXPENSES:', f"KES {total_expenses:,.2f}"])
            table_data.append(['', '', '', 'NET PROFIT:', f"KES {total_revenue - total_expenses:,.2f}"])
                
        elif report_type == 'payments':
            headers = ['Payment Date', 'Order #', 'Customer', 'Method', 'Amount']
            table_data = [headers]
            total_amount = 0
            for item in items[:50]:
                amount = float(item.amount or 0)
                total_amount += amount
                table_data.append([
                    item.completed_at.strftime('%Y-%m-%d %H:%M') if item.completed_at else 'N/A',
                    item.service_order.order_number if item.service_order else 'N/A',
                    item.service_order.customer.full_name if item.service_order and item.service_order.customer else 'N/A',
                    item.payment_method.name if item.payment_method else 'Cash',
                    f"KES {amount:,.2f}"
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_amount:,.2f}"])
        
        elif report_type == 'daily_summary':
            headers = ['Date', 'Orders Count', 'Revenue', 'Paid Orders', 'Order Value']
            table_data = [headers]
            total_revenue = 0
            total_orders = 0
            for item in items[:50]:
                revenue = float(item.get('revenue_from_payments', 0) or 0)
                orders_count = int(item.get('orders_count', 0) or 0)
                paid_orders = int(item.get('paid_orders_count', 0) or 0)
                order_value = float(item.get('total_order_value', 0) or 0)
                
                total_revenue += revenue
                total_orders += orders_count
                
                # Handle date field - could be 'created_at__date' or 'date'
                date_str = item.get('created_at__date') or item.get('date', 'N/A')
                if hasattr(date_str, 'strftime'):
                    date_str = date_str.strftime('%Y-%m-%d')
                
                table_data.append([
                    str(date_str),
                    str(orders_count),
                    f"KES {revenue:,.2f}",
                    str(paid_orders),
                    f"KES {order_value:,.2f}"
                ])
            # Add total row
            table_data.append(['TOTAL:', str(total_orders), f"KES {total_revenue:,.2f}", '', ''])
                
        elif report_type == 'weekly_summary':
            headers = ['Week', 'Orders Count', 'Revenue', 'Completed Orders']
            table_data = [headers]
            total_revenue = 0
            for item in items[:50]:
                revenue = float(item.get('revenue_from_payments', 0) or 0)
                total_revenue += revenue
                week_str = item.get('week', 'N/A')
                if hasattr(week_str, 'strftime'):
                    week_str = week_str.strftime('%Y-W%U')
                table_data.append([
                    str(week_str),
                    str(item.get('orders_count', 0) or 0),
                    f"KES {revenue:,.2f}",
                    str(item.get('paid_orders_count', 0) or 0)
                ])
            # Add total row
            table_data.append(['', 'TOTAL:', f"KES {total_revenue:,.2f}", ''])
                
        elif report_type == 'monthly_summary':
            headers = ['Month', 'Orders Count', 'Revenue', 'Completed Orders']
            table_data = [headers]
            total_revenue = 0
            for item in items[:50]:
                revenue = float(item.get('revenue_from_payments', 0) or 0)
                total_revenue += revenue
                month_str = item.get('month', 'N/A')
                if hasattr(month_str, 'strftime'):
                    month_str = month_str.strftime('%Y-%m')
                table_data.append([
                    str(month_str),
                    str(item.get('orders_count', 0) or 0),
                    f"KES {revenue:,.2f}",
                    str(item.get('paid_orders_count', 0) or 0)
                ])
            # Add total row
            table_data.append(['', 'TOTAL:', f"KES {total_revenue:,.2f}", ''])
                
        elif report_type == 'sales_analysis':
            headers = ['Service', 'Category', 'Quantity', 'Revenue', 'Orders']
            table_data = [headers]
            total_revenue = 0
            for item in items[:50]:
                revenue = float(item.get('total_revenue', 0) or 0)
                total_revenue += revenue
                service_name = item.get('service__name', 'N/A')
                category_name = item.get('service__category__name', 'Uncategorized')
                quantity = item.get('quantity_sold', 0) or 0
                orders_count = item.get('orders_count', 0) or 0
                table_data.append([
                    str(service_name),
                    str(category_name),
                    str(quantity),
                    f"KES {revenue:,.2f}",
                    str(orders_count)
                ])
            # Add total row
            table_data.append(['', '', '', f"KES {total_revenue:,.2f}", ''])
                
        elif report_type == 'expenses':
            headers = ['Date', 'Description', 'Category', 'Vendor', 'Amount']
            table_data = [headers]
            total_amount = 0
            for item in items[:50]:
                amount = float(item.total_amount or 0)
                total_amount += amount
                table_data.append([
                    item.expense_date.strftime('%Y-%m-%d'),
                    item.description or 'N/A',
                    item.category.name if item.category else 'N/A',
                    item.vendor.name if item.vendor else 'N/A',
                    f"KES {amount:,.2f}"
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_amount:,.2f}"])
                
        elif report_type == 'vehicle_analysis':
            headers = ['Vehicle', 'Customer', 'License Plate', 'Services', 'Total Spent', 'Last Service']
            table_data = [headers]
            total_revenue = 0
            total_services = 0
            for item in items[:50]:
                spent = float(item.total_spent or 0)
                services = int(item.paid_services_count or 0)
                total_revenue += spent
                total_services += services
                
                vehicle_name = f"{item.make or ''} {item.model or ''}".strip() or 'Unknown Vehicle'
                customer_name = item.customer.full_name if item.customer else 'No Customer'
                license_plate = item.license_plate or 'No Plate'
                last_service = item.last_service.strftime('%Y-%m-%d') if item.last_service else 'Never'
                
                table_data.append([
                    vehicle_name,
                    customer_name,
                    license_plate,
                    str(services),
                    f"KES {spent:,.2f}",
                    last_service
                ])
            # Add total row
            table_data.append(['', '', 'TOTALS:', str(total_services), f"KES {total_revenue:,.2f}", ''])
        
        elif report_type == 'balance_sheet':
            headers = ['Category', 'Item', 'Amount', 'Type']
            table_data = [headers]
            
            # Summary section first
            summary = report_data.get('summary', {})
            table_data.append(['SUMMARY', 'Total Assets', f"KES {summary.get('total_assets', 0):,.2f}", 'Asset'])
            table_data.append(['SUMMARY', 'Total Liabilities', f"KES {summary.get('total_liabilities', 0):,.2f}", 'Liability'])
            table_data.append(['SUMMARY', 'Total Equity', f"KES {summary.get('total_equity', 0):,.2f}", 'Equity'])
            table_data.append(['', '', '', ''])  # Spacing
            
            # Detailed items
            for item in items[:50]:
                table_data.append([
                    item.get('category', ''),
                    item.get('name', ''),
                    f"KES {item.get('amount', 0):,.2f}",
                    item.get('type', '').title()
                ])
            
            # Financial performance
            table_data.append(['', '', '', ''])  # Spacing
            table_data.append(['PERFORMANCE', 'Total Revenue', f"KES {summary.get('total_revenue', 0):,.2f}", 'Income'])
            table_data.append(['PERFORMANCE', 'Total Expenses', f"KES {summary.get('total_expenses', 0):,.2f}", 'Expense'])
            table_data.append(['PERFORMANCE', 'Net Income', f"KES {summary.get('net_income', 0):,.2f}", 'Profit/Loss'])
        
        elif report_type == 'refunds_report':
            headers = ['Refund ID', 'Type', 'Order', 'Customer', 'Amount', 'Date']
            table_data = [headers]
            total_refund_amount = 0
            
            for item in items[:50]:  # Limit to 50 items for PDF
                if isinstance(item, dict):
                    amount = float(item.get('amount', 0))
                    total_refund_amount += amount
                    processed_at = item.get('processed_at')
                    date_str = processed_at.strftime('%Y-%m-%d') if processed_at else 'N/A'
                    
                    table_data.append([
                        item.get('refund_id', ''),
                        item.get('type', ''),
                        item.get('original_order', ''),
                        item.get('customer', ''),
                        f"KES {amount:,.2f}",
                        date_str
                    ])
                else:
                    # Fallback for unexpected item format
                    table_data.append([
                        str(item)[:20],
                        'Refund Item',
                        'N/A',
                        'N/A',
                        'N/A',
                        'N/A'
                    ])
            
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_refund_amount:,.2f}", ''])
        
        else:
            # Generic format for other report types
            headers = ['Item', 'Details', 'Value']
            table_data = [headers]
            for item in items[:50]:
                table_data.append([
                    str(item)[:40],
                    f"{report_type.title()} data",
                    'N/A'
                ])
        
        return table_data
    
    def _generate_excel_export(self, report_data, tenant):
        """Generate comprehensive Excel for specific report type"""
        if not EXCEL_AVAILABLE:
            return JsonResponse({'error': 'Excel generation not available'}, status=500)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_data.get('title', 'Report')
        
        # Header with tenant information
        ws['A1'] = tenant.name
        ws['A1'].font = Font(bold=True, size=16)
        ws['B1'] = f"Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
        ws['B1'].font = Font(size=10)
        
        ws['A2'] = getattr(tenant, 'address', 'Address not available')
        ws['B2'] = f"Report Period: {report_data.get('period', '')}"
        ws['B2'].font = Font(bold=True)
        
        # Main title
        current_row = 4
        ws[f'A{current_row}'] = report_data.get('title', 'Business Report')
        ws[f'A{current_row}'].font = Font(bold=True, size=14)
        current_row += 2
        
        # Add employee details for individual employee report
        if report_data.get('type') == 'individual_employee' and report_data.get('employee_details'):
            employee_details = report_data['employee_details']
            ws[f'A{current_row}'] = 'EMPLOYEE INFORMATION'
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            
            # Employee details headers
            ws[f'A{current_row}'] = 'Field'
            ws[f'B{current_row}'] = 'Value'
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # Employee details data
            employee_fields = [
                ('Employee ID', employee_details.get('employee_id', 'N/A')),
                ('Full Name', employee_details.get('full_name', 'N/A')),
                ('Email', employee_details.get('email', 'N/A')),
                ('Phone', employee_details.get('phone_number', 'N/A')),
                ('Department', employee_details.get('department', 'N/A')),
                ('Position', str(employee_details.get('position', 'N/A'))),
                ('Hire Date', str(employee_details.get('hire_date', 'N/A'))),
                ('Status', 'Active' if employee_details.get('is_active') else 'Inactive')
            ]
            
            for field_name, field_value in employee_fields:
                ws[f'A{current_row}'] = field_name
                ws[f'B{current_row}'] = field_value
                current_row += 1
            
            current_row += 2  # Add spacing
        
        # Summary section
        if report_data.get('summary'):
            ws[f'A{current_row}'] = 'SUMMARY'
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            
            # Summary headers
            ws[f'A{current_row}'] = 'Metric'
            ws[f'B{current_row}'] = 'Value'
            ws[f'A{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].font = Font(bold=True)
            current_row += 1
            
            for key, value in report_data['summary'].items():
                if key != 'error' and value is not None:
                    formatted_key = key.replace('_', ' ').title()
                    ws[f'A{current_row}'] = formatted_key
                    if 'revenue' in key.lower() or 'profit' in key.lower() or 'value' in key.lower() or 'amount' in key.lower():
                        ws[f'B{current_row}'] = f"KES {value:,.2f}"
                    elif 'rate' in key.lower() or 'margin' in key.lower():
                        ws[f'B{current_row}'] = f"{value:.1f}%" if isinstance(value, (int, float)) else str(value)
                    else:
                        ws[f'B{current_row}'] = str(value)
                    current_row += 1
            
            current_row += 2  # Add spacing
        
        # Detailed Data section
        if report_data.get('items'):
            ws[f'A{current_row}'] = 'DETAILED DATA'
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            
            # Generate Excel table data
            table_data = self._generate_excel_table_data(report_data)
            
            if table_data:
                # Add headers
                for col_idx, header in enumerate(table_data[0], 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                current_row += 1
                
                # Add data rows
                for row_data in table_data[1:]:
                    for col_idx, cell_value in enumerate(row_data, 1):
                        ws.cell(row=current_row, column=col_idx, value=cell_value)
                    current_row += 1
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Footer
        current_row += 2
        ws[f'A{current_row}'] = f"Report generated by {tenant.name} - AutoWash Management System"
        ws[f'A{current_row}'].font = Font(italic=True, size=9)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{tenant.name}_{report_data.get("type", "report")}_report.xlsx"'
        return response
    
    def _generate_excel_table_data(self, report_data):
        """Generate table data for Excel based on report type - Updated for payment-based data"""
        report_type = report_data.get('type')
        items = report_data.get('items', [])
        
        if not items:
            return []
        
        if report_type == 'business_overview':
            headers = ['Payment Date', 'Order #', 'Customer', 'Payment Method', 'Amount', 'Status']
            table_data = [headers]
            total_amount = 0
            for item in items:
                amount = float(item.amount or 0)
                total_amount += amount
                table_data.append([
                    item.completed_at.strftime('%Y-%m-%d') if item.completed_at else 'N/A',
                    item.service_order.order_number if item.service_order else 'N/A',
                    item.service_order.customer.full_name if item.service_order and item.service_order.customer else 'N/A',
                    item.payment_method.name if item.payment_method else 'Cash',
                    amount,
                    item.status.title()
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', total_amount, ''])
                
        elif report_type == 'inventory':
            headers = ['Item Name', 'Category', 'Current Stock', 'Unit Price', 'Stock Value', 'Movements']
            table_data = [headers]
            total_value = 0
            for item in items:
                unit_price = float(getattr(item, 'unit_cost', 0) or 0)
                stock_value = float(item.current_stock or 0) * unit_price
                total_value += stock_value
                table_data.append([
                    item.name,
                    item.category.name if item.category else 'N/A',
                    int(item.current_stock or 0),
                    unit_price,
                    stock_value,
                    int(getattr(item, 'movement_count', 0) or 0)
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL VALUE:', total_value, ''])
                
        elif report_type == 'services':
            headers = ['Service Name', 'Category', 'Base Price', 'Payments Count', 'Total Revenue', 'Avg Rating']
            table_data = [headers]
            total_revenue = 0
            for item in items:
                revenue = float(getattr(item, 'total_revenue', 0) or 0)
                total_revenue += revenue
                table_data.append([
                    item.name,
                    item.category.name if item.category else 'N/A',
                    float(item.base_price or 0),
                    int(getattr(item, 'payments_count', 0) or 0),
                    revenue,
                    round(float(getattr(item, 'avg_rating', 0) or 0), 2)
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', total_revenue, ''])
                
        elif report_type == 'customers' or report_type == 'customer_analytics':
            headers = ['Customer Name', 'Email', 'Phone', 'Total Payments', 'Total Spent', 'Avg Payment Value']
            table_data = [headers]
            total_spent = 0
            for item in items:
                spent = float(getattr(item, 'total_spent', 0) or 0)
                total_spent += spent
                table_data.append([
                    item.full_name,
                    item.email or 'N/A',
                    str(getattr(item, 'phone', 'N/A')) or 'N/A',
                    int(getattr(item, 'total_payments', 0) or 0),
                    spent,
                    float(getattr(item, 'avg_payment_value', 0) or 0)
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', total_spent, ''])
                
        elif report_type == 'employees':
            headers = ['Employee Name', 'Department', 'Position', 'Payments Handled', 'Revenue Generated', 'Avg Payment Value']
            table_data = [headers]
            total_revenue = 0
            for item in items:
                revenue = float(getattr(item, 'revenue_generated', 0) or 0)
                total_revenue += revenue
                table_data.append([
                    item.full_name,
                    item.department.name if item.department else 'N/A',
                    item.position.title if item.position else 'N/A',
                    int(getattr(item, 'payments_handled', 0) or 0),
                    revenue,
                    float(getattr(item, 'avg_payment_value', 0) or 0)
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', total_revenue, ''])
                
        elif report_type == 'individual_employee':
            employee = report_data.get('employee')
            if employee:
                headers = ['Payment Date', 'Order Number', 'Customer', 'Payment Method', 'Amount', 'Status']
                table_data = [headers]
                total_amount = 0
                for item in items:
                    amount = float(item.amount or 0)
                    total_amount += amount
                    table_data.append([
                        item.completed_at.strftime('%Y-%m-%d') if item.completed_at else 'N/A',
                        item.service_order.order_number if item.service_order else 'N/A',
                        item.service_order.customer.full_name if item.service_order and item.service_order.customer else 'N/A',
                        item.payment_method.name if item.payment_method else 'Cash',
                        amount,
                        item.status.title()
                    ])
                # Add total row
                table_data.append(['', '', '', 'TOTAL:', total_amount, ''])
            else:
                return []
                
        elif report_type == 'financial_summary':
            headers = ['Date', 'Transaction Type', 'Description', 'Category', 'Amount']
            table_data = [headers]
            total_revenue = 0
            total_expenses = 0
            for item in items:
                amount = float(item.get('amount', 0) or 0)
                if item.get('type') == 'Revenue':
                    total_revenue += amount
                else:
                    total_expenses += abs(amount)
                table_data.append([
                    item.get('date', '').strftime('%Y-%m-%d') if hasattr(item.get('date', ''), 'strftime') else str(item.get('date', '')),
                    item.get('type', 'N/A'),
                    item.get('description', 'N/A'),
                    item.get('category', 'N/A'),
                    amount
                ])
            # Add summary rows
            table_data.append(['', '', '', 'TOTAL REVENUE:', total_revenue])
            table_data.append(['', '', '', 'TOTAL EXPENSES:', total_expenses])
            table_data.append(['', '', '', 'NET PROFIT:', total_revenue - total_expenses])
                
        elif report_type == 'payments':
            headers = ['Payment Date & Time', 'Order Number', 'Customer', 'Payment Method', 'Amount', 'Status']
            table_data = [headers]
            total_amount = 0
            for item in items:
                amount = float(item.amount or 0)
                total_amount += amount
                table_data.append([
                    item.completed_at.strftime('%Y-%m-%d %H:%M') if item.completed_at else 'N/A',
                    item.service_order.order_number if item.service_order else 'N/A',
                    item.service_order.customer.full_name if item.service_order and item.service_order.customer else 'N/A',
                    item.payment_method.name if item.payment_method else 'Cash',
                    amount,
                    item.status.title()
                ])
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', total_amount, ''])
                
        elif report_type == 'expenses':
            headers = ['Date', 'Description', 'Category', 'Vendor', 'Amount', 'Reference']
            table_data = [headers]
            total_amount = 0
            for item in items:
                amount = float(item.total_amount or 0)
                total_amount += amount
                table_data.append([
                    item.expense_date.strftime('%Y-%m-%d'),
                    item.description or 'N/A',
                    item.category.name if item.category else 'N/A',
                    item.vendor.name if item.vendor else 'N/A',
                    amount,
                    getattr(item, 'reference_number', 'N/A')
                ])
            # Add total row
            table_data.append(['', '', '', '', 'TOTAL:', total_amount])
            
        elif report_type == 'daily_summary':
            headers = ['Date', 'Payments', 'Revenue', 'Expenses', 'Net Profit', 'Customers']
            table_data = [headers]
            total_revenue = 0
            total_expenses = 0
            total_payments = 0
            total_customers = 0
            for item in items:
                revenue = float(item.get('revenue', 0) or 0)
                expenses = float(item.get('expenses', 0) or 0)
                payments = int(item.get('payments_count', 0) or 0)
                customers = int(item.get('customers', 0) or 0)
                total_revenue += revenue
                total_expenses += expenses
                total_payments += payments
                total_customers += customers
                table_data.append([
                    item.get('date', '').strftime('%Y-%m-%d') if hasattr(item.get('date', ''), 'strftime') else str(item.get('date', '')),
                    payments,
                    revenue,
                    expenses,
                    revenue - expenses,
                    customers
                ])
            # Add total row
            table_data.append(['TOTAL:', total_payments, total_revenue, total_expenses, total_revenue - total_expenses, total_customers])
            
        elif report_type == 'weekly_summary':
            headers = ['Week Starting', 'Orders', 'Revenue', 'Expenses', 'Net Profit', 'Customers']
            table_data = [headers]
            total_revenue = 0
            total_expenses = 0
            total_orders = 0
            total_customers = 0
            for item in items:
                revenue = float(item.get('revenue_from_payments', 0) or 0)
                expenses = float(item.get('expenses', 0) or 0)
                orders = int(item.get('orders_count', 0) or 0)
                customers = int(item.get('unique_customers', 0) or 0)
                total_revenue += revenue
                total_expenses += expenses
                total_orders += orders
                total_customers += customers
                table_data.append([
                    item.get('week', '').strftime('%Y-%m-%d') if hasattr(item.get('week', ''), 'strftime') else str(item.get('week', '')),
                    orders,
                    revenue,
                    expenses,
                    revenue - expenses,
                    customers
                ])
            # Add total row
            table_data.append(['TOTAL:', total_orders, total_revenue, total_expenses, total_revenue - total_expenses, total_customers])
            
        elif report_type == 'monthly_summary':
            headers = ['Month', 'Orders', 'Revenue', 'Expenses', 'Net Profit', 'Customers']
            table_data = [headers]
            total_revenue = 0
            total_expenses = 0
            total_orders = 0
            total_customers = 0
            for item in items:
                revenue = float(item.get('revenue_from_payments', 0) or 0)
                expenses = float(item.get('expenses', 0) or 0)
                orders = int(item.get('orders_count', 0) or 0)
                customers = int(item.get('unique_customers', 0) or 0)
                total_revenue += revenue
                total_expenses += expenses
                total_orders += orders
                total_customers += customers
                table_data.append([
                    item.get('month', '').strftime('%Y-%m') if hasattr(item.get('month', ''), 'strftime') else str(item.get('month', '')),
                    orders,
                    revenue,
                    expenses,
                    revenue - expenses,
                    customers
                ])
            # Add total row
            table_data.append(['TOTAL:', total_orders, total_revenue, total_expenses, total_revenue - total_expenses, total_customers])
            
        elif report_type == 'sales_analysis':
            headers = ['Service', 'Category', 'Quantity Sold', 'Revenue', 'Order Value', 'Orders Count']
            table_data = [headers]
            total_revenue = 0
            total_order_value = 0
            total_orders = 0
            for item in items:
                revenue = float(item.get('total_revenue', 0) or 0)
                order_value = float(item.get('total_order_value', 0) or 0)
                orders = int(item.get('orders_count', 0) or 0)
                quantity = int(item.get('quantity_sold', 0) or 0)
                service_name = item.get('service__name', 'N/A')
                category_name = item.get('service__category__name', 'Uncategorized')
                
                total_revenue += revenue
                total_order_value += order_value
                total_orders += orders
                
                table_data.append([
                    service_name,
                    category_name,
                    quantity,
                    revenue,
                    order_value,
                    orders
                ])
            # Add total row
            table_data.append(['TOTAL:', '', '', total_revenue, total_order_value, total_orders])
                
        elif report_type == 'vehicle_analysis':
            headers = ['Vehicle', 'Customer', 'License Plate', 'Services Count', 'Total Spent', 'Last Service', 'Customer Phone']
            table_data = [headers]
            total_revenue = 0
            total_services = 0
            for item in items:
                spent = float(item.total_spent or 0)
                services = int(item.paid_services_count or 0)
                total_revenue += spent
                total_services += services
                
                vehicle_name = f"{item.make or ''} {item.model or ''}".strip() or 'Unknown Vehicle'
                customer_name = item.customer.full_name if item.customer else 'No Customer'
                customer_phone = str(item.customer.phone) if item.customer and item.customer.phone else 'N/A'
                license_plate = item.license_plate or 'No Plate'
                last_service = item.last_service.strftime('%Y-%m-%d') if item.last_service else 'Never'
                
                table_data.append([
                    vehicle_name,
                    customer_name,
                    license_plate,
                    services,
                    spent,
                    last_service,
                    customer_phone
                ])
            # Add total row
            table_data.append(['', '', 'TOTALS:', total_services, total_revenue, '', ''])
        
        elif report_type == 'balance_sheet':
            headers = ['Category', 'Item', 'Amount', 'Type']
            table_data = [headers]
            
            # Process balance sheet items
            for item in items:
                if isinstance(item, dict):
                    table_data.append([
                        item.get('category', ''),
                        item.get('name', ''),
                        f"KES {float(item.get('amount', 0)):,.2f}",
                        item.get('type', '').title()
                    ])
                else:
                    # Fallback for unexpected item format
                    table_data.append([
                        str(item)[:50],
                        'Balance Sheet Item',
                        'N/A',
                        'Unknown'
                    ])
        
        elif report_type == 'refunds_report':
            headers = ['Refund ID', 'Type', 'Order', 'Customer', 'Amount', 'Reason', 'Processed By', 'Date']
            table_data = [headers]
            total_refund_amount = 0
            
            # Process refund items
            for item in items:
                if isinstance(item, dict):
                    amount = float(item.get('amount', 0))
                    total_refund_amount += amount
                    processed_at = item.get('processed_at')
                    date_str = processed_at.strftime('%Y-%m-%d') if processed_at else 'N/A'
                    
                    table_data.append([
                        item.get('refund_id', ''),
                        item.get('type', ''),
                        item.get('original_order', ''),
                        item.get('customer', ''),
                        f"KES {amount:,.2f}",
                        item.get('reason', '')[:50] + ('...' if len(item.get('reason', '')) > 50 else ''),
                        item.get('processed_by', ''),
                        date_str
                    ])
                else:
                    # Fallback for unexpected item format
                    table_data.append([
                        str(item)[:20],
                        'Refund Item',
                        'N/A',
                        'N/A',
                        'N/A',
                        'N/A',
                        'N/A',
                        'N/A'
                    ])
            
            # Add total row
            table_data.append(['', '', '', 'TOTAL:', f"KES {total_refund_amount:,.2f}", '', '', ''])
        
        else:
            # Generic format for other report types
            headers = ['Item', 'Description', 'Value', 'Date']
            table_data = [headers]
            for item in items:
                table_data.append([
                    str(item)[:50],
                    f"{report_type.title()} data",
                    getattr(item, 'amount', getattr(item, 'total_amount', 0)) or 0,
                    getattr(item, 'created_at', timezone.now()).strftime('%Y-%m-%d')
                ])
        
        return table_data
